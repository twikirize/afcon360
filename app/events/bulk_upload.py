"""
Bulk event registration via Excel/CSV upload.

Two supported upload modes:
  - external_group       -> public authenticated user / organisation
                            registering attendees on their behalf.
                            Paid events require immediate payment.
  - coordinator_managed  -> authorized event organizer/coordinator.
                            May create hosted/comped/invoice registrations.

Only the official downloaded template is accepted.
"""
from flask import Blueprint, request, render_template, jsonify, session, current_app, send_file
from flask_login import login_required, current_user
from app.extensions import db, redis_client
from app.events.constants import MAX_BULK_GROUP_SIZE
from app.events.services import EventService, SoldOutException
from app.events.models import Event, TicketType, EventRegistration
from app.events.permissions import can_manage_event
import io
import uuid
import json
import logging
import secrets
import time

logger = logging.getLogger(__name__)

bulk_bp = Blueprint('event_bulk', __name__, url_prefix='/events/bulk')


# ============================================================================
# HELPERS
# ============================================================================

def _store_bulk_upload(attendees, errors=None, ttl_seconds=3600):
    """
    Store parsed bulk attendees outside the Flask session cookie.

    Redis is preferred. Fallback to Flask session exists only for local
    development and small lists; production should use Redis or a staging DB.
    """
    payload = {
        "attendees": attendees,
        "errors": errors or [],
        "created_at": time.time(),
    }

    if redis_client:
        upload_id = secrets.token_urlsafe(24)
        try:
            redis_client.setex(
                f"bulk:upload:{upload_id}",
                ttl_seconds,
                json.dumps(payload),
            )
            return upload_id
        except Exception as exc:
            logger.warning("Redis bulk upload storage failed: %s", exc)

    # Fallback: small session list, with warning.
    logger.warning(
        "Using Flask session for bulk upload payload. This is not safe for large lists."
    )
    session["bulk_attendees"] = attendees
    session["bulk_errors"] = errors or []
    return "session"


def _load_bulk_upload(upload_id):
    """Return attendees/errors for a previously stored bulk upload."""
    if upload_id == "session":
        return session.get("bulk_attendees", []), session.get("bulk_errors", [])

    if not redis_client:
        return [], []

    raw = redis_client.get(f"bulk:upload:{upload_id}")
    if not raw:
        return [], []

    try:
        if isinstance(raw, bytes):
            raw = raw.decode("utf-8")
        payload = json.loads(raw)
        return payload.get("attendees", []), payload.get("errors", [])
    except Exception as exc:
        logger.error("Failed to decode bulk upload %s: %s", upload_id, exc)
        return [], []


def _clear_bulk_upload(upload_id):
    if upload_id == "session":
        session.pop("bulk_attendees", None)
        session.pop("bulk_errors", None)
        return

    if redis_client:
        try:
            redis_client.delete(f"bulk:upload:{upload_id}")
        except Exception as exc:
            logger.warning("Failed to clear bulk upload key: %s", exc)


def _validate_event_availability(event, ticket_type_id):
    """Reject bulk registration if the event/ticket type is not open."""
    gate_error = EventService._registration_gate_error(event, ticket_type_id)
    return gate_error


def _ticket_remaining_capacity(ticket_type):
    """Return remaining seats correctly.

    `available_seats` is remaining capacity when set. Unlimited capacity is
    represented by `capacity == 0`.
    """
    if ticket_type.capacity and ticket_type.capacity > 0:
        return ticket_type.available_seats if ticket_type.available_seats is not None else ticket_type.capacity
    return None


# ============================================================================
# TEMPLATE
# ============================================================================

@bulk_bp.route('/<identifier>/template')
@login_required
def download_template(identifier):
    """Download Excel template for bulk registration"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendees"

    # Headers
    headers = ['Full Name', 'Email Address', 'Phone Number', 'Nationality']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

    # Example row
    ws.cell(row=2, column=1, value="John Doe")
    ws.cell(row=2, column=2, value="john@example.com")
    ws.cell(row=2, column=3, value="+256 700 123 456")
    ws.cell(row=2, column=4, value="Ugandan")

    # Adjust column widths
    for col in range(1, 5):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'attendee_template_{identifier}.xlsx'
    )


# ============================================================================
# UPLOAD / VALIDATE
# ============================================================================

@bulk_bp.route('/<identifier>/upload', methods=['POST'])
@login_required
def upload_bulk(identifier):
    """
    Upload and validate a bulk registration file.

    Request may include:
      - file
      - bulk_mode: external_group | coordinator_managed
    """
    import pandas as pd

    try:
        event = Event.query.filter_by(slug=identifier).first()
        if not event:
            return jsonify({'success': False, 'error': 'Event not found'}), 404

        bulk_mode = (
            request.form.get('bulk_mode')
            or request.args.get('bulk_mode')
            or 'external_group'
        )

        if bulk_mode not in {'external_group', 'coordinator_managed'}:
            return jsonify({'success': False, 'error': f'Invalid bulk_mode: {bulk_mode}'}), 400

        # Coordinator-managed bulk uploads require event management authority.
        if bulk_mode == 'coordinator_managed':
            allowed, permission_error = can_manage_event(current_user, event)
            if not allowed:
                return jsonify({
                    'success': False,
                    'error': permission_error or 'Unauthorized to manage this event',
                }), 403

        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400

        # Parse file using only the official template format.
        if file.filename.endswith('.xlsx'):
            df = pd.read_excel(file)
        elif file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            return jsonify({'success': False, 'error': 'Unsupported file type. Use .xlsx or .csv'}), 400

        expected_columns = ['Full Name', 'Email Address', 'Phone Number', 'Nationality']
        missing_columns = [col for col in expected_columns if col not in df.columns]
        if missing_columns:
            return jsonify({
                'success': False,
                'error': f'Missing columns: {", ".join(missing_columns)}'
            }), 400

        attendees = []
        errors = []

        for idx, row in df.iterrows():
            name = str(row.get('Full Name', '')).strip()
            email = str(row.get('Email Address', '')).strip().lower()
            phone = str(row.get('Phone Number', '')).strip()
            nationality = str(row.get('Nationality', '')).strip()

            if not name or not email:
                errors.append(f"Row {idx + 2}: Missing name or email")
                continue

            if '@' not in email or '.' not in email:
                errors.append(f"Row {idx + 2}: Invalid email address")
                continue

            attendees.append({
                'name': name,
                'email': email,
                'phone': phone if phone and phone != 'nan' else None,
                'nationality': nationality if nationality and nationality != 'nan' else None,
            })

        if len(attendees) > MAX_BULK_GROUP_SIZE:
            return jsonify({
                'success': False,
                'error': f'File contains {len(attendees)} attendees. Maximum is {MAX_BULK_GROUP_SIZE}.'
            }), 400

        upload_id = _store_bulk_upload(attendees, errors)

        return jsonify({
            'success': True,
            'upload_id': upload_id,
            'bulk_mode': bulk_mode,
            'total': len(attendees),
            'errors': errors,
            'preview': attendees[:10],
        })

    except Exception as e:
        logger.error(f"Bulk upload error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# CONFIRM / PROCESS
# ============================================================================

@bulk_bp.route('/<identifier>/confirm', methods=['POST'])
@login_required
def confirm_bulk(identifier):
    """
    Confirm and process a bulk registration.

    Request JSON:
      {
        "upload_id": "...",
        "ticket_type_id": 123,
        "group_label": "Company delegation",
        "payment_method": "wallet",          # external paid events
        "mobile_money_operator": "mtn",      # optional external paid
        "mobile_money_phone": "07...",       # optional external paid
        "bulk_mode": "external_group",       # or coordinator_managed
        "payment_mode": "hosted"             # coordinator-managed paid only:
                                             # hosted | comped | invoice
      }
    """
    try:
        data = request.get_json(silent=True) or {}
        upload_id = data.get('upload_id') or 'session'
        bulk_mode = data.get('bulk_mode') or 'external_group'

        event = Event.query.filter_by(slug=identifier).first()
        if not event:
            return jsonify({'success': False, 'error': 'Event not found'}), 404

        attendees, upload_errors = _load_bulk_upload(upload_id)
        if not attendees:
            return jsonify({'success': False, 'error': 'No attendees to register'}), 400

        ticket_type_id = data.get('ticket_type_id')
        group_label = data.get('group_label', '').strip() or None

        ticket_type = TicketType.query.filter_by(
            id=ticket_type_id,
            event_id=event.id,
        ).first()
        if not ticket_type:
            return jsonify({'success': False, 'error': 'Invalid ticket type'}), 400

        # Validate mode and permission again at confirmation.
        if bulk_mode not in {'external_group', 'coordinator_managed'}:
            return jsonify({'success': False, 'error': f'Invalid bulk_mode: {bulk_mode}'}), 400

        if bulk_mode == 'coordinator_managed':
            allowed, permission_error = can_manage_event(current_user, event)
            if not allowed:
                return jsonify({
                    'success': False,
                    'error': permission_error or 'Unauthorized to manage this event',
                }), 403

        # Event / ticket availability gate.
        gate_error = _validate_event_availability(event, ticket_type_id)
        if gate_error:
            return jsonify({'success': False, 'error': gate_error}), 410

        # Capacity check.
        remaining = _ticket_remaining_capacity(ticket_type)
        if remaining is not None and remaining < len(attendees):
            return jsonify({
                'success': False,
                'error': f'Only {remaining} tickets available, but {len(attendees)} requested',
            }), 400

        # Generate one group ID for the entire batch.
        group_booking_id = str(uuid.uuid4())

        registrations = []
        errors = list(upload_errors or [])

        # ────────────────────────────────────────────────────────────────
        # EXTERNAL GROUP BOOKER
        # ────────────────────────────────────────────────────────────────
        if bulk_mode == 'external_group':
            is_paid = bool(ticket_type.price and float(ticket_type.price) > 0)

            if is_paid:
                # Paid external bulk bookings must pay through the existing
                # payment service before any registrations are persisted.
                from app.events.payment_service import EventPaymentService

                payment_service = EventPaymentService()
                payment_result = payment_service.process_ticket_purchase(
                    user_id=current_user.id,
                    event_id=event.id,
                    ticket_type_id=ticket_type_id,
                    quantity=len(attendees),
                    payment_method=data.get('payment_method', 'wallet'),
                    mobile_money_operator=data.get('mobile_money_operator'),
                    mobile_money_phone=data.get('mobile_money_phone'),
                    group_attendees=attendees,
                    create_primary_for_payer=False,
                    group_booking_id=group_booking_id,
                )

                if not payment_result.get('success'):
                    return jsonify({
                        'success': False,
                        'error': payment_result.get('error') or 'Payment failed',
                    }), 400

                registrations = payment_result.get('registrations') or []
            else:
                # Free external bulk registration.
                for idx, attendee in enumerate(attendees, start=1):
                    registration_data = {
                        'ticket_type_id': ticket_type_id,
                        'full_name': attendee['name'],
                        'email': attendee['email'],
                        'phone': attendee.get('phone', ''),
                        'nationality': attendee.get('nationality', ''),
                    }

                    reg, qr, err = EventService.register_for_event(
                        identifier,
                        current_user.id,
                        registration_data,
                        booking_type="third_party",
                        attendee_email=attendee['email'],
                        attendee_name=attendee['name'],
                        attendee_phone=attendee.get('phone'),
                        group_booking_id=group_booking_id,
                        group_index=idx,
                        group_label=group_label,
                    )

                    if err:
                        errors.append(f"{attendee['name']}: {err}")
                    else:
                        registrations.append(reg)

        # ────────────────────────────────────────────────────────────────
        # COORDINATOR-MANAGED BULK
        # ────────────────────────────────────────────────────────────────
        else:
            payment_mode = data.get('payment_mode', 'hosted')
            if payment_mode not in {'hosted', 'comped', 'invoice'}:
                return jsonify({
                    'success': False,
                    'error': f'Invalid coordinator payment_mode: {payment_mode}',
                }), 400

            for idx, attendee in enumerate(attendees, start=1):
                registration_data = {
                    'ticket_type_id': ticket_type_id,
                    'full_name': attendee['name'],
                    'email': attendee['email'],
                    'phone': attendee.get('phone', ''),
                    'nationality': attendee.get('nationality', ''),
                }

                reg, qr, err = EventService.register_for_event(
                    identifier,
                    current_user.id,
                    registration_data,
                    booking_type="third_party",
                    attendee_email=attendee['email'],
                    attendee_name=attendee['name'],
                    attendee_phone=attendee.get('phone'),
                    group_booking_id=group_booking_id,
                    group_index=idx,
                    group_label=group_label,
                )

                if err:
                    errors.append(f"{attendee['name']}: {err}")
                    continue

                # Coordinator-managed paid registrations need non-wallet status.
                # The service initially marks paid ticket registrations as
                # pending payment, so update them here for hosted/comped/invoice.
                if ticket_type.price and float(ticket_type.price) > 0:
                    EventRegistration.query.filter_by(
                        registration_ref=reg['registration_ref']
                    ).update({
                        'payment_status': payment_mode,
                        'status': 'confirmed',
                    })

                registrations.append(reg)

        # Commit any coordinator payment_status updates.
        try:
            db.session.commit()
        except Exception as commit_error:
            db.session.rollback()
            logger.error("Bulk confirm commit error: %s", commit_error)
            return jsonify({'success': False, 'error': 'Failed to finalize bulk registration'}), 500

        _clear_bulk_upload(upload_id)

        if not registrations:
            return jsonify({
                'success': False,
                'error': f'Failed to register any attendees: {"; ".join(errors)}',
            }), 400

        first_ref = registrations[0].get('registration_ref')
        if not first_ref:
            return jsonify({'success': False, 'error': 'Registration reference missing'}), 500

        first_reg = EventRegistration.query.filter_by(registration_ref=first_ref).first()

        if first_reg:
            qr_code = EventService._generate_qr_code(
                first_reg.qr_token,
                first_reg.registration_ref,
            )
            session['last_registration'] = {
                'registration': EventService._registration_to_dict(first_reg),
                'qr_code': qr_code,
                'event': EventService.get_event(identifier),
                'group_registrations': registrations,
                'errors': errors,
            }

        return jsonify({
            'success': True,
            'registered': len(registrations),
            'errors': errors,
            'redirect': f'/events/registration-confirmation/{first_ref}',
        })

    except Exception as e:
        db.session.rollback()
        logger.error(f"Bulk confirm error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500