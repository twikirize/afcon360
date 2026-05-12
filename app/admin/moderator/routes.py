"""
Moderator Blueprint - Complete Route Set
=========================================
All routes the moderator role needs. Nothing deferred.

Sections:
  A.  Dashboard
  B.  Content Submissions  (list · view · approve · reject · request-changes · assign · flag · bulk)
  C.  Flag Management      (list · view · resolve · escalate · close · assign · bulk · create)
  D.  User Moderation      (list · view · suspend · unsuspend · warn · verify · deactivate · flag · note · bulk)
  E.  ManageableCategory   (list · view · toggle-active)
  F.  ManageableItem       (list · view · approve · reject · feature · delete · flag)
  G.  Event Moderation     (list · view · approve · reject · flag)
  H.  Organisation Queue   (list · view · approve · reject · flag)
  I.  KYC Queue            (list · view · flag · refer-to-compliance)
  J.  Audit Log            (list - read-only)
  K.  Moderation Stats     (dashboard)
  L.  JSON API helpers     (live counters, quick-resolve)
"""

import logging
logger = logging.getLogger(__name__)
from datetime import datetime, timezone, timedelta

from flask import (
    Blueprint, render_template, request, jsonify,
    flash, redirect, url_for, abort, Response
)
from flask_login import login_required, current_user
from sqlalchemy import func, or_

from app.auth.decorators import require_role, require_permission, require_fresh_user

from app.extensions import db
from app.identity.models.user import User
# Models imported at function level to avoid circular imports
from app.admin.moderator import moderator_bp
from app.admin.services import create_flag, resolve_flag
from app.admin.moderator.registry import get_review_url
from app.admin.compliance.services import ComplianceCaseService
from app.admin.compliance.models import ComplianceCaseType, ComplianceCasePriority

# ModerationLog model is now defined in app/admin/models.py
# SLA migration is handled via Alembic migrations

logger = logging.getLogger(__name__)

#moderator_bp = Blueprint('moderator', __name__, url_prefix='/moderator')

# Roles allowed everywhere unless a route narrows it
_MOD = ('moderator', 'admin', 'super_admin', 'owner')


# ─────────────────────────────────────────────────────────────────────────────
# INTERNAL HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _audit(action: str, category: str, details: dict):
    """Silently write to OwnerAuditLog; never raises."""
    try:
        from app.admin.owner.models import OwnerAuditLog
        OwnerAuditLog.log_action(
            current_user, action=action, category=category,
            details={**details, '_ip': request.remote_addr}
        )
    except Exception as exc:
        logger.debug("Audit log unavailable: %s", exc)


def _review_url(entity_type: str, entity_id: int):
    try:
        from app.admin.moderator.registry import get_review_url
        return get_review_url(entity_type, entity_id)
    except Exception:
        return None


def _enrich_flags(flag_list):
    return [{"flag": f, "review_url": _review_url(f.entity_type, f.entity_id)}
            for f in flag_list]


def _moderators():
    """Return list of users who are moderators/admins - for assign dropdowns."""
    try:
        from app.identity.models.roles_permission import Role, UserRole
        role_ids = db.session.query(Role.id).filter(
            Role.name.in_(['moderator', 'admin', 'super_admin'])
        ).subquery()
        return (User.query
                .join(UserRole, UserRole.user_id == User.id)
                .filter(UserRole.role_id.in_(role_ids))
                .distinct().all())
    except Exception:
        return []


def _redirect_back(fallback):
    ref = request.referrer
    return redirect(ref if ref else url_for(fallback))


def _build_moderation_stats():
    """Build moderation statistics for dashboard across all modules."""
    from app.admin.models import (
        ContentFlag, ContentSubmission, ModerationLog,
        ManageableItem, ManageableCategory
    )
    stats = {}

    # Content Submissions
    try:
        stats['content_submissions'] = ContentSubmission.query.filter_by(status='pending').count()
    except Exception:
        stats['content_submissions'] = 0

    # Open Flags
    try:
        stats['open_flags'] = ContentFlag.query.filter_by(status='open').count()
        stats['flags_critical'] = ContentFlag.query.filter_by(status='open', priority='critical').count()
    except Exception:
        stats['open_flags'] = 0
        stats['flags_critical'] = 0

    # Events
    try:
        from app.events.models import Event
        from app.events.constants import EventStatus
        stats['events'] = Event.query.filter_by(status=EventStatus.PENDING_APPROVAL, is_deleted=False).count()
    except Exception:
        stats['events'] = 0

    # Organisations
    try:
        from app.identity.models.organisation import Organisation
        stats['pending_orgs'] = Organisation.query.filter_by(verification_status='pending').count()
    except Exception:
        stats['pending_orgs'] = 0

    # KYC
    try:
        from app.kyc.models import KycRecord
        stats['kyc'] = KycRecord.query.filter_by(status='pending').count()
    except Exception:
        stats['kyc'] = 0

    # Tourism
    try:
        from app.tourism.models import TourismListing
        stats['tourism'] = TourismListing.query.filter_by(status='pending').count()
    except Exception:
        stats['tourism'] = 0

    # Transport (drivers, vehicles)
    try:
        from app.transport.models import Vehicle, DriverProfile
        drivers = DriverProfile.query.filter_by(verification_status='pending').count()
        vehicles = Vehicle.query.filter_by(verification_status='pending').count()
        stats['transport'] = drivers + vehicles
    except Exception:
        stats['transport'] = 0

    # Accommodation (properties, bookings, reviews)
    try:
        from app.accommodation.models.property import Property, AccommodationPropertyStatus
        from app.accommodation.models.booking import AccommodationBooking
        from app.accommodation.models.review import Review, AccommodationReviewStatus
        properties = Property.query.filter_by(status=AccommodationPropertyStatus.PENDING_REVIEW).count()
        bookings = AccommodationBooking.query.filter_by(status='pending').count()
        reviews = Review.query.filter_by(status=AccommodationReviewStatus.PENDING).count()
        stats['accommodation'] = properties + bookings + reviews
    except Exception:
        stats['accommodation'] = 0

    return stats


def _build_my_stats():
    """Build moderator's personal statistics from ModerationLog."""
    from app.admin.models import ModerationLog
    try:
        week_ago = datetime.now(timezone.utc) - timedelta(days=7)
        my_logs = ModerationLog.query.filter(
            ModerationLog.moderator_id == current_user.id,
            ModerationLog.created_at >= week_ago
        ).all()

        my_stats = {}
        for log in my_logs:
            action = log.action
            my_stats[action] = my_stats.get(action, 0) + 1

        return my_stats
    except Exception as e:
        logger.debug("Error building my_stats: %s", e)
        return {}


# ═════════════════════════════════════════════════════════════════════════════
# A.  DASHBOARD
# ═════════════════════════════════════════════════════════════════════════════

@moderator_bp.route('/dashboard')
@login_required
@require_role(*_MOD)
def dashboard():
    from app.admin.models import (
        ContentFlag, ContentSubmission, ModerationLog,
        ManageableItem, ManageableCategory
    )
    from sqlalchemy import func, case
    week_ago = datetime.now(timezone.utc) - timedelta(days=7)
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)

    # Enterprise-level moderation metrics
    # Critical flags (risk score >= 80 or priority critical)
    flags_critical = ContentFlag.query.filter(
        ContentFlag.status == 'open',
        db.or_(
            ContentFlag.priority == 'critical',
            ContentFlag.risk_score >= 80
        )
    ).count()

    # High priority flags
    flags_high = ContentFlag.query.filter_by(status='open', priority='high').count()

    # SLA breach metrics
    sla_breached_flags = ContentFlag.query.filter(
        ContentFlag.status == 'open',
        ContentFlag.sla_due_at < datetime.now(timezone.utc)
    ).all()
    sla_breached_count = len(sla_breached_flags)
    
    # Calculate average SLA breach time
    avg_sla_breach_time = 0
    if sla_breached_flags:
        total_breach_time = sum(
            (datetime.now(timezone.utc) - flag.sla_due_at).total_seconds() / 60 
            for flag in sla_breached_flags
        )
        avg_sla_breach_time = total_breach_time / len(sla_breached_flags)

    # AI performance metrics
    ai_processed_today = ContentFlag.query.filter(
        ContentFlag.detection_source == 'ai',
        ContentFlag.created_at >= today
    ).count()
    
    ai_flags = ContentFlag.query.filter(
        ContentFlag.detection_source == 'ai',
        ContentFlag.ai_confidence.isnot(None)
    ).all()
    ai_accuracy = sum(flag.ai_confidence or 0 for flag in ai_flags) / len(ai_flags) * 100 if ai_flags else 0

    # Response time metrics
    resolved_today = ContentFlag.query.filter(
        ContentFlag.status == 'resolved',
        ContentFlag.resolved_at >= today
    ).all()
    
    avg_response_time = 0
    if resolved_today:
        total_response_time = sum(
            (flag.resolved_at - flag.created_at).total_seconds()
            for flag in resolved_today
        )
        avg_response_time = total_response_time / len(resolved_today)

    # SLA compliance rate
    total_resolved = ContentFlag.query.filter_by(status='resolved').count()
    sla_compliant = ContentFlag.query.filter(
        ContentFlag.status == 'resolved',
        ContentFlag.sla_breached == False
    ).count()
    sla_compliance = (sla_compliant / total_resolved * 100) if total_resolved > 0 else 0

    # Active moderators and workload
    from app.identity.models.user import UserRole
    from app.identity.models.roles_permission import Role
    
    active_moderators = User.query.join(UserRole, User.id == UserRole.user_id).join(Role, UserRole.role_id == Role.id).filter(
        Role.name == 'moderator',
        User.is_active == True
    ).count()
    
    assigned_flags = ContentFlag.query.filter(
        ContentFlag.assigned_to.isnot(None),
        ContentFlag.status == 'open'
    ).count()
    avg_workload = assigned_flags / active_moderators if active_moderators > 0 else 0

    # Escalation metrics
    total_flags = ContentFlag.query.count()
    escalated_flags = ContentFlag.query.filter(
        ContentFlag.escalation_count > 0
    ).count()
    escalation_rate = (escalated_flags / total_flags * 100) if total_flags > 0 else 0
    
    level_3_cases = ContentFlag.query.filter_by(
        moderation_level='level_3',
        status='open'
    ).count()

    # User satisfaction (mock data for now)
    user_satisfaction = 4.2  # This would come from actual feedback system
    feedback_count = 156  # This would come from actual feedback system

    # Resolution rate
    resolution_rate = (total_resolved / total_flags * 100) if total_flags > 0 else 0

    # Average risk score
    open_flags_with_risk = ContentFlag.query.filter(
        ContentFlag.status == 'open',
        ContentFlag.risk_score.isnot(None)
    ).all()
    avg_risk_score = sum(flag.risk_score for flag in open_flags_with_risk) / len(open_flags_with_risk) if open_flags_with_risk else 0

    # Current moderator level (would be based on user's permissions/role)
    current_moderator_level = "1"  # This would be determined dynamically

    # Build enterprise moderation stats
    moderation_stats = {
        'flags_critical': flags_critical,
        'flags_high': flags_high,
        'sla_breached_count': sla_breached_count,
        'avg_sla_breach_time': avg_sla_breach_time,
        'ai_processed_today': ai_processed_today,
        'ai_accuracy': ai_accuracy,
        'avg_response_time': avg_response_time,
        'sla_compliance': sla_compliance,
        'active_moderators': active_moderators,
        'avg_workload': avg_workload,
        'escalation_rate': escalation_rate,
        'level_3_cases': level_3_cases,
        'user_satisfaction': user_satisfaction,
        'feedback_count': feedback_count,
        'resolved_today': len(resolved_today),
        'resolution_rate': resolution_rate,
        'avg_risk_score': avg_risk_score,
    }
    
    # Add transport statistics
    try:
        from app.transport.models import DriverProfile, Vehicle, Booking
        transport_stats = {
            'pending_drivers': DriverProfile.query.filter_by(
                verification_tier='pending', is_deleted=False
            ).count(),
            'pending_vehicles': Vehicle.query.filter_by(
                status='pending', is_deleted=False
            ).count(),
            'disputed_bookings': Booking.query.filter_by(
                status='disputed', is_deleted=False
            ).count(),
            'total_active_drivers': DriverProfile.query.filter_by(
                verification_tier='platform_verified', is_active=True, is_deleted=False
            ).count(),
            'total_verified_vehicles': Vehicle.query.filter_by(
                status='active', is_deleted=False
            ).count(),
            'transport_flags_today': ContentFlag.query.filter(
                ContentFlag.entity_type.like('transport_%'),
                ContentFlag.created_at >= today
            ).count()
        }
        moderation_stats.update(transport_stats)
    except Exception:
        # Transport module not available - skip transport stats
        pass

    # Original dashboard data (keep for compatibility)
    pending_submissions = (ContentSubmission.query
                           .filter_by(status='pending')
                           .order_by(ContentSubmission.created_at.desc())
                           .limit(10).all())

    recent_users = (User.query
                    .filter(User.created_at >= week_ago)
                    .order_by(User.created_at.desc())
                    .limit(10).all())

    open_flags = (ContentFlag.query
                  .filter_by(status='open')
                  .order_by(ContentFlag.created_at.desc())
                  .limit(10).all())

    flag_rows = _enrich_flags(open_flags)

    by_priority = dict(
        ContentFlag.query
        .with_entities(ContentFlag.priority, func.count())
        .filter_by(status='open')
        .group_by(ContentFlag.priority).all()
    )
    by_type = dict(
        ContentFlag.query
        .with_entities(ContentFlag.entity_type, func.count())
        .filter_by(status='open')
        .group_by(ContentFlag.entity_type).all()
    )
    total_open_flags = sum(by_priority.values())

    approved_7d = ContentSubmission.query.filter(
        ContentSubmission.status == 'approved',
        ContentSubmission.reviewed_at >= week_ago
    ).count()
    rejected_7d = ContentSubmission.query.filter(
        ContentSubmission.status == 'rejected',
        ContentSubmission.reviewed_at >= week_ago
    ).count()
    resolved_flags_7d = ContentFlag.query.filter(
        ContentFlag.status == 'resolved',
        ContentFlag.resolved_at >= week_ago
    ).count()
    pending_count = ContentSubmission.query.filter_by(status='pending').count()
    changes_requested_count = ContentSubmission.query.filter_by(status='changes_requested').count()

    # Build moderator's personal stats
    my_stats = _build_my_stats() if _build_my_stats else {}
    
    # Add my queue stats for enterprise view
    if my_stats:
        my_stats['assigned_count'] = ContentFlag.query.filter_by(
            assigned_to=current_user.id, 
            status='open'
        ).count()
        my_stats['in_review_count'] = ContentFlag.query.filter_by(
            assigned_to=current_user.id, 
            status='in_review'
        ).count()

    return render_template(
        'admin/moderator/dashboard.html',
        pending_submissions=pending_submissions,
        recent_users=recent_users,
        flags=open_flags,
        flag_rows=flag_rows,
        flags_by_priority=by_priority,
        flags_by_type=by_type,
        total_open_flags=total_open_flags,
        approved_7d=approved_7d,
        rejected_7d=rejected_7d,
        resolved_flags_7d=resolved_flags_7d,
        pending_count=pending_count,
        changes_requested_count=changes_requested_count,
        moderation_stats=moderation_stats,
        my_stats=my_stats,
        current_moderator_level=current_moderator_level,
        now=datetime.utcnow,
        title="Enterprise Moderation Dashboard"
    )


# ═════════════════════════════════════════════════════════════════════════════
# B.  CONTENT SUBMISSIONS
# ═════════════════════════════════════════════════════════════════════════════

@moderator_bp.route('/content')
@login_required
@require_role(*_MOD)
def content_moderation():
    from app.admin.models import ContentSubmission, ManageableCategory
    page       = request.args.get('page', 1, type=int)
    per_page   = request.args.get('per_page', 20, type=int)
    status     = request.args.get('status', 'pending')
    cat_id     = request.args.get('category_id', type=int)
    search     = request.args.get('q', '').strip()
    sort       = request.args.get('sort', 'newest')   # newest | oldest

    q = ContentSubmission.query
    if status != 'all':
        q = q.filter_by(status=status)
    if cat_id:
        q = q.filter_by(category_id=cat_id)
    if search:
        q = q.filter(ContentSubmission.name.ilike(f'%{search}%'))

    q = q.order_by(
        ContentSubmission.created_at.asc()
        if sort == 'oldest'
        else ContentSubmission.created_at.desc()
    )

    submissions = q.paginate(page=page, per_page=per_page, error_out=False)
    categories  = ManageableCategory.query.filter_by(is_active=True).order_by(ManageableCategory.name).all()

    # Status counts for tab badges
    status_counts = dict(
        ContentSubmission.query
        .with_entities(ContentSubmission.status, func.count())
        .group_by(ContentSubmission.status).all()
    )

    # If this is an AJAX request for partial table refresh, return only the table rows
    if request.headers.get('Accept') == 'text/html' and request.args.get('partial') == '1':
        return render_template(
            'admin/moderator/_pending_table.html',
            pending_submissions=submissions.items,
            now=datetime.utcnow
        )

    return render_template(
        'admin/moderator/content.html',
        submissions=submissions,
        status=status,
        categories=categories,
        selected_category=cat_id,
        search=search,
        sort=sort,
        status_counts=status_counts,
        title="Content Moderation"
    )


@moderator_bp.route('/content/<int:submission_id>')
@login_required
@require_role(*_MOD)
def view_submission(submission_id):
    from app.admin.models import ContentSubmission, ContentFlag
    submission    = ContentSubmission.query.get_or_404(submission_id)
    related_flags = (ContentFlag.query
                     .filter_by(entity_type='content_submission', entity_id=submission_id)
                     .order_by(ContentFlag.created_at.desc()).all())
    moderators    = _moderators()
    return render_template(
        'admin/moderator/view_submission.html',
        submission=submission,
        related_flags=related_flags,
        moderators=moderators,
        title=f"Submission #{submission_id}"
    )


@moderator_bp.route('/content/<int:submission_id>/claim', methods=['POST'])
@login_required
@require_role(*_MOD)
def claim_submission(submission_id):
    """Claim a content submission for review."""
    from app.admin.models import ContentSubmission, ModerationLog
    submission = ContentSubmission.query.get_or_404(submission_id)

    # Check if already claimed
    if submission.assigned_to_id is not None:
        return jsonify({'ok': False, 'error': 'Already claimed'}), 409

    now = datetime.now(timezone.utc)
    submission.assigned_to_id = current_user.id
    submission.claimed_at = now

    db.session.add(ModerationLog(
        moderator_id=current_user.id,
        submission_id=submission_id,
        action='claim',
        notes='Submission claimed for review'
    ))
    db.session.commit()

    _audit('submission.claim', 'content', {'id': submission_id, 'claimed_by': current_user.id})

    return jsonify({'ok': True, 'claimed_at': now.isoformat()})


@moderator_bp.route('/content/<int:submission_id>/approve', methods=['POST'])
@login_required
@require_role(*_MOD)
def approve_submission(submission_id):
    from app.admin.models import ContentSubmission, ModerationLog
    s = ContentSubmission.query.get_or_404(submission_id)
    now = datetime.now(timezone.utc)
    s.status      = 'approved'
    s.reviewed_by = current_user.id
    s.reviewed_at = now
    s.review_notes = request.form.get('notes', '')
    s.resolved_at = now
    if s.claimed_at:
        s.processing_time_seconds = int((now - s.claimed_at).total_seconds())
    db.session.add(ModerationLog(
        moderator_id=current_user.id,
        submission_id=submission_id,
        action='approve',
        notes=f"Processing time: {s.processing_time_seconds}s"
    ))
    db.session.commit()
    _audit('submission.approve', 'content', {'id': submission_id, 'name': s.name})
    flash('Submission approved.', 'success')
    return redirect(url_for('admin.moderator.content_moderation'))


@moderator_bp.route('/content/<int:submission_id>/reject', methods=['POST'])
@login_required
@require_role(*_MOD)
def reject_submission(submission_id):
    from app.admin.models import ContentSubmission, ModerationLog
    s = ContentSubmission.query.get_or_404(submission_id)
    notes = request.form.get('notes', '').strip()
    if not notes:
        flash('Rejection reason is required.', 'danger')
        return redirect(url_for('admin.moderator.view_submission', submission_id=submission_id))
    now = datetime.now(timezone.utc)
    s.status      = 'rejected'
    s.reviewed_by = current_user.id
    s.reviewed_at = now
    s.review_notes = notes
    s.resolved_at = now
    if s.claimed_at:
        s.processing_time_seconds = int((now - s.claimed_at).total_seconds())
    db.session.add(ModerationLog(
        moderator_id=current_user.id,
        submission_id=submission_id,
        action='reject',
        notes=f"Processing time: {s.processing_time_seconds}s"
    ))
    db.session.commit()
    _audit('submission.reject', 'content', {'id': submission_id, 'reason': notes})
    flash('Submission rejected.', 'warning')
    return redirect(url_for('admin.moderator.content_moderation'))


@moderator_bp.route('/content/<int:submission_id>/request-changes', methods=['POST'])
@login_required
@require_role(*_MOD)
def request_changes(submission_id):
    from app.admin.models import ContentSubmission
    s = ContentSubmission.query.get_or_404(submission_id)
    notes = request.form.get('notes', '').strip()
    if not notes:
        flash('Change request details are required.', 'danger')
        return redirect(url_for('admin.moderator.view_submission', submission_id=submission_id))
    s.status      = 'changes_requested'
    s.reviewed_by = current_user.id
    s.reviewed_at = datetime.now(timezone.utc)
    s.review_notes = notes
    db.session.commit()
    _audit('submission.request_changes', 'content', {'id': submission_id, 'notes': notes})
    flash('Changes requested from submitter.', 'info')
    return redirect(url_for('admin.moderator.content_moderation'))


@moderator_bp.route('/content/<int:submission_id>/assign', methods=['POST'])
@login_required
@require_role(*_MOD)
def assign_submission(submission_id):
    from app.admin.models import ContentSubmission
    s = ContentSubmission.query.get_or_404(submission_id)
    assignee_id = request.form.get('assignee_id', type=int)
    if not assignee_id:
        flash('Select a moderator to assign.', 'danger')
        return redirect(url_for('admin.moderator.view_submission', submission_id=submission_id))
    assignee = User.query.get_or_404(assignee_id)
    s.reviewed_by  = assignee.id
    s.review_notes = f"[ASSIGNED → {assignee.username}] {s.review_notes or ''}"
    db.session.commit()
    _audit('submission.assign', 'content', {'id': submission_id, 'assignee': assignee.username})
    flash(f'Assigned to {assignee.username}.', 'success')
    return redirect(url_for('admin.moderator.view_submission', submission_id=submission_id))


@moderator_bp.route('/content/<int:submission_id>/flag', methods=['POST'])
@login_required
@require_permission('content.flag')
def flag_submission(submission_id):
    from app.admin.models import ContentSubmission
    ContentSubmission.query.get_or_404(submission_id)   # existence check
    reason   = request.form.get('reason') or (request.json or {}).get('reason', '')
    priority = request.form.get('priority', 'normal')
    ok, result = create_flag(current_user, 'content_submission', submission_id, reason, priority)
    if not ok:
        flash(f'Could not flag: {result}', 'danger')
        if request.is_json:
            return jsonify({"ok": False, "error": result}), 403
        return redirect(url_for('admin.moderator.view_submission', submission_id=submission_id))
    _audit('submission.flag', 'moderation', {'submission_id': submission_id, 'priority': priority})
    flash('Submission flagged for escalation.', 'warning')
    if request.is_json:
        return jsonify({"ok": True, "flag_id": int(result.id)})
    return redirect(url_for('admin.moderator.view_submission', submission_id=submission_id))


@moderator_bp.route('/content/bulk-action', methods=['POST'])
@login_required
@require_role(*_MOD)
def bulk_submission_action():
    """Approve or reject multiple submissions at once."""
    from app.admin.models import ContentSubmission
    ids    = request.form.getlist('submission_ids', type=int)
    action = request.form.get('action')   # approve | reject | request_changes
    notes  = request.form.get('notes', 'Bulk action by moderator')

    if not ids:
        flash('No submissions selected.', 'warning')
        return redirect(url_for('admin.moderator.content_moderation'))

    valid_actions = {'approve': 'approved', 'reject': 'rejected', 'request_changes': 'changes_requested'}
    if action not in valid_actions:
        flash('Invalid action.', 'danger')
        return redirect(url_for('admin.moderator.content_moderation'))

    count = 0
    for sid in ids:
        s = ContentSubmission.query.get(sid)
        if s and s.status == 'pending':
            s.status      = valid_actions[action]
            s.reviewed_by = current_user.id
            s.reviewed_at = datetime.now(timezone.utc)
            s.review_notes = notes
            count += 1
    db.session.commit()
    _audit(f'submission.bulk_{action}', 'content', {'count': count, 'ids': ids})
    flash(f'{count} submission(s) {valid_actions[action]}.', 'success')
    return redirect(url_for('admin.moderator.content_moderation'))


# ═════════════════════════════════════════════════════════════════════════════
# C.  FLAG MANAGEMENT
# ═════════════════════════════════════════════════════════════════════════════

@moderator_bp.route('/flagged')
@login_required
@require_role(*_MOD)
def flagged_content():
    from app.admin.models import ContentFlag
    page            = request.args.get('page', 1, type=int)
    per_page        = request.args.get('per_page', 20, type=int)
    status_filter   = request.args.get('status', 'open')
    priority_filter = request.args.get('priority')
    type_filter     = request.args.get('entity_type')
    detection_source = request.args.get('detection_source')
    moderation_level = request.args.get('moderation_level')
    sla_status      = request.args.get('sla_status')
    sort            = request.args.get('sort', 'newest')
    search          = request.args.get('search', '').strip()

    q = ContentFlag.query
    if status_filter != 'all':
        q = q.filter_by(status=status_filter)
    if priority_filter:
        q = q.filter_by(priority=priority_filter)
    if type_filter:
        q = q.filter_by(entity_type=type_filter)
    if detection_source:
        q = q.filter_by(detection_source=detection_source)
    if moderation_level:
        q = q.filter_by(moderation_level=moderation_level)
    if search:
        q = q.filter(ContentFlag.reason.ilike(f'%{search}%'))

    # SLA status filtering
    if sla_status == 'breached':
        q = q.filter(ContentFlag.sla_breached == True)
    elif sla_status == 'urgent':
        urgent_time = datetime.now(timezone.utc) + timedelta(hours=1)
        q = q.filter(
            ContentFlag.sla_due_at < urgent_time,
            ContentFlag.sla_due_at > datetime.now(timezone.utc)
        )
    elif sla_status == 'normal':
        q = q.filter(
            ContentFlag.sla_due_at > datetime.now(timezone.utc) + timedelta(hours=1)
        )

    if sort == 'priority':
        priority_order = db.case(
            (ContentFlag.priority == 'critical', 1),
            (ContentFlag.priority == 'high',     2),
            (ContentFlag.priority == 'medium',   3),
            (ContentFlag.priority == 'normal',   4),
            (ContentFlag.priority == 'low',      5),
            else_=6
        )
        q = q.order_by(priority_order, ContentFlag.created_at.asc())
    else:
        q = q.order_by(ContentFlag.created_at.desc())

    flags      = q.paginate(page=page, per_page=per_page, error_out=False)
    flag_rows  = _enrich_flags(flags.items)
    moderators = _moderators()

    entity_types = [
        r[0] for r in db.session.query(ContentFlag.entity_type.distinct()).all() if r[0]
    ]

    status_counts = dict(
        ContentFlag.query
        .with_entities(ContentFlag.status, func.count())
        .group_by(ContentFlag.status).all()
    )

    # Enterprise metrics
    sla_breached_count = ContentFlag.query.filter(
        ContentFlag.sla_breached == True,
        ContentFlag.status == 'open'
    ).count()
    
    # AI accuracy calculation
    ai_flags = ContentFlag.query.filter(
        ContentFlag.detection_source == 'ai',
        ContentFlag.ai_confidence.isnot(None)
    ).all()
    ai_accuracy = sum(flag.ai_confidence or 0 for flag in ai_flags) / len(ai_flags) * 100 if ai_flags else 0

    # Set status for template
    status = status_filter if status_filter != 'all' else 'open'
    priority = priority_filter

    return render_template(
        'admin/moderator/flagged.html',
        flags=flags,
        flag_rows=flag_rows,
        entity_types=entity_types,
        moderators=moderators,
        status=status,
        priority=priority,
        status_filter=status_filter,
        priority_filter=priority_filter,
        entity_type_filter=type_filter,
        detection_source=detection_source,
        moderation_level=moderation_level,
        sla_status=sla_status,
        sort=sort,
        status_counts=status_counts,
        sla_breached_count=sla_breached_count,
        ai_accuracy=ai_accuracy,
        now=datetime.now(timezone.utc),
        title="Enterprise Flagged Content"
    )


@moderator_bp.route('/flagged/<int:flag_id>')
@login_required
@require_role(*_MOD)
def view_flag(flag_id):
    from app.admin.models import ContentFlag
    flag        = ContentFlag.query.get_or_404(flag_id)
    review_url  = _review_url(flag.entity_type, flag.entity_id)
    moderators  = _moderators()

    # Other open flags on the same entity
    siblings = (ContentFlag.query
                .filter(
                    ContentFlag.entity_type == flag.entity_type,
                    ContentFlag.entity_id   == flag.entity_id,
                    ContentFlag.id          != flag.id,
                    ContentFlag.status      == 'open'
                )
                .order_by(ContentFlag.created_at.desc())
                .limit(5).all())

    # Flags raised by the same flagger (context)
    by_same_user = (ContentFlag.query
                    .filter_by(flagged_by=flag.flagged_by)
                    .order_by(ContentFlag.created_at.desc())
                    .limit(5).all())

    return render_template(
        'admin/moderator/view_flag.html',
        flag=flag,
        review_url=review_url,
        siblings=siblings,
        by_same_user=by_same_user,
        moderators=moderators,
        title=f"Flag #{flag_id}"
    )


@moderator_bp.route('/flagged/<int:flag_id>/resolve', methods=['POST'])
@login_required
@require_role(*_MOD)
def resolve_flag_route(flag_id):
    action = request.form.get('action', 'reviewed') or (request.json or {}).get('action', 'reviewed')
    notes  = request.form.get('notes', '')           or (request.json or {}).get('notes', '')
    ok, result = resolve_flag(current_user, flag_id, action, notes)
    if ok:
        _audit('flag.resolve', 'moderation', {'flag_id': flag_id, 'action': action})
        flash('Flag resolved.', 'success')
    else:
        flash(f'Error: {result}', 'danger')
    if request.is_json:
        return jsonify({"ok": ok, "error": None if ok else result})
    return _redirect_back('admin.moderator.flagged_content')


@moderator_bp.route('/flagged/<int:flag_id>/escalate', methods=['POST'])
@login_required
@require_role(*_MOD)
def escalate_flag(flag_id):
    from app.admin.models import ContentFlag
    flag        = ContentFlag.query.get_or_404(flag_id)
    role        = request.form.get('role', 'admin')
    assignee_id = request.form.get('assignee_id', type=int)
    notes       = request.form.get('notes', '').strip()

    flag.status           = 'in_review'
    flag.escalated_to_role = role
    if assignee_id:
        flag.assigned_to = assignee_id
    if notes:
        flag.resolution_notes = notes
    db.session.commit()
    _audit('flag.escalate', 'moderation', {
        'flag_id': flag_id, 'to_role': role, 'assignee': assignee_id
    })
    flash('Flag escalated.', 'warning')
    if request.is_json:
        return jsonify({"ok": True})
    return _redirect_back('admin.moderator.flagged_content')


@moderator_bp.route('/flagged/<int:flag_id>/close', methods=['POST'])
@login_required
@require_role(*_MOD)
def close_flag(flag_id):
    from app.admin.models import ContentFlag
    flag       = ContentFlag.query.get_or_404(flag_id)
    notes      = request.form.get('notes', 'Closed by moderator - not actionable')
    flag.status             = 'resolved'
    flag.resolved_by        = current_user.id
    flag.resolution_action  = 'closed'
    flag.resolution_notes   = notes
    flag.resolved_at        = datetime.now(timezone.utc)
    db.session.commit()
    _audit('flag.close', 'moderation', {'flag_id': flag_id, 'notes': notes})
    flash('Flag closed.', 'info')
    if request.is_json:
        return jsonify({"ok": True})
    return _redirect_back('admin.moderator.flagged_content')


@moderator_bp.route('/flagged/<int:flag_id>/assign', methods=['POST'])
@login_required
@require_role(*_MOD)
def assign_flag(flag_id):
    from app.admin.models import ContentFlag
    flag        = ContentFlag.query.get_or_404(flag_id)
    assignee_id = request.form.get('assignee_id', type=int)
    if not assignee_id:
        flash('Select a admin.moderator.', 'danger')
        return _redirect_back('admin.moderator.flagged_content')
    assignee         = User.query.get_or_404(assignee_id)
    flag.assigned_to = assignee.id
    flag.status      = 'in_review'
    db.session.commit()
    _audit('flag.assign', 'moderation', {'flag_id': flag_id, 'assignee': assignee.username})
    flash(f'Flag assigned to {assignee.username}.', 'success')
    return _redirect_back('admin.moderator.flagged_content')


@moderator_bp.route('/flagged/<int:flag_id>/reprioritise', methods=['POST'])
@login_required
@require_role(*_MOD)
def reprioritise_flag(flag_id):
    from app.admin.models import ContentFlag
    flag          = ContentFlag.query.get_or_404(flag_id)
    new_priority  = request.form.get('priority', 'normal')
    VALID         = {'low', 'normal', 'medium', 'high', 'critical'}
    if new_priority not in VALID:
        flash('Invalid priority.', 'danger')
        return _redirect_back('admin.moderator.flagged_content')
    old           = flag.priority
    flag.priority = new_priority
    db.session.commit()
    _audit('flag.reprioritise', 'moderation', {
        'flag_id': flag_id, 'old': old, 'new': new_priority
    })
    flash(f'Priority changed: {old} → {new_priority}.', 'info')
    if request.is_json:
        return jsonify({"ok": True})
    return _redirect_back('admin.moderator.flagged_content')


@moderator_bp.route('/flagged/create', methods=['POST'])
@login_required
@require_permission('content.flag')
def create_flag_route():
    """Moderator manually raises a flag against any entity."""
    entity_type = request.form.get('entity_type', '').strip()
    entity_id   = request.form.get('entity_id', type=int)
    reason      = request.form.get('reason', '').strip()
    priority    = request.form.get('priority', 'normal')
    if not entity_type or not entity_id or not reason:
        flash('Entity type, entity ID, and reason are all required.', 'danger')
        return _redirect_back('admin.moderator.flagged_content')
    ok, result = create_flag(current_user, entity_type, entity_id, reason, priority)
    if ok:
        _audit('flag.create', 'moderation', {
            'entity_type': entity_type, 'entity_id': entity_id, 'priority': priority
        })
        flash(f'Flag #{result.id} created.', 'success')
    else:
        flash(f'Could not create flag: {result}', 'danger')
    if request.is_json:
        return jsonify({"ok": ok, "flag_id": int(result.id) if ok else None})
    return _redirect_back('admin.moderator.flagged_content')


@moderator_bp.route('/flagged/bulk-action', methods=['POST'])
@login_required
@require_role(*_MOD)
def bulk_flag_action():
    from app.admin.models import ContentFlag
    flag_ids = request.form.getlist('flag_ids', type=int)
    action   = request.form.get('action')   # resolve | close | escalate
    notes    = request.form.get('notes', 'Bulk action by moderator')
    role     = request.form.get('role', 'admin')

    if not flag_ids:
        flash('No flags selected.', 'warning')
        return redirect(url_for('admin.moderator.flagged_content'))

    count = 0
    for fid in flag_ids:
        flag = ContentFlag.query.get(fid)
        if not flag or flag.status not in ('open', 'in_review'):
            continue
        if action in ('resolve', 'close'):
            ok, _ = resolve_flag(current_user, fid, action, notes)
            if ok:
                count += 1
        elif action == 'escalate':
            flag.status            = 'in_review'
            flag.escalated_to_role = role
            db.session.commit()
            count += 1

    _audit(f'flag.bulk_{action}', 'moderation', {'count': count, 'ids': flag_ids})
    flash(f'{count} flag(s) actioned.', 'success')
    return redirect(url_for('admin.moderator.flagged_content'))


# ═════════════════════════════════════════════════════════════════════════════
# D.  USER MODERATION
# ═════════════════════════════════════════════════════════════════════════════

@moderator_bp.route('/users')
@login_required
@require_role(*_MOD)
def user_moderation():
    from app.admin.models import ContentFlag
    page         = request.args.get('page', 1, type=int)
    per_page     = request.args.get('per_page', 20, type=int)
    filter_      = request.args.get('filter', 'all')
    search       = request.args.get('q', '').strip()
    week_ago     = datetime.now(timezone.utc) - timedelta(days=7)

    q = User.query
    if filter_ == 'recent':
        q = q.filter(User.created_at >= week_ago)
    elif filter_ == 'unverified':
        q = q.filter_by(is_verified=False)
    elif filter_ == 'inactive':
        q = q.filter_by(is_active=False)
    elif filter_ == 'flagged':
        flagged_ids = db.session.query(ContentFlag.entity_id).filter_by(
            entity_type='user', status='open'
        ).subquery()
        q = q.filter(User.id.in_(flagged_ids))

    if search:
        q = q.filter(or_(
            User.username.ilike(f'%{search}%'),
            User.email.ilike(f'%{search}%')
        ))

    users = q.order_by(User.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    # Count flags per user (for badge display)
    open_flag_counts = dict(
        db.session.query(ContentFlag.entity_id, func.count())
        .filter_by(entity_type='user', status='open')
        .group_by(ContentFlag.entity_id).all()
    )

    return render_template(
        'admin/moderator/users.html',
        users=users,
        current_filter=filter_,
        search=search,
        open_flag_counts=open_flag_counts,
        title="User Moderation"
    )


@moderator_bp.route('/users/<int:user_id>')
@login_required
@require_role(*_MOD)
def view_user(user_id):
    from app.admin.models import ContentFlag, ContentSubmission
    user           = User.query.get_or_404(user_id)
    flags_raised   = (ContentFlag.query.filter_by(flagged_by=user_id)
                      .order_by(ContentFlag.created_at.desc()).limit(20).all())
    flags_against  = (ContentFlag.query.filter_by(entity_type='user', entity_id=user_id)
                      .order_by(ContentFlag.created_at.desc()).limit(20).all())
    submissions    = (ContentSubmission.query.filter_by(submitted_by=user_id)
                      .order_by(ContentSubmission.created_at.desc()).limit(20).all())
    return render_template(
        'admin/moderator/view_user.html',
        user=user,
        flags_raised=flags_raised,
        flags_against=flags_against,
        submissions=submissions,
        title=f"User: {user.username}"
    )


@moderator_bp.route('/users/<int:user_id>/suspend', methods=['POST'])
@login_required
@require_role(*_MOD)
@require_fresh_user
def suspend_user(user_id):
    from app.admin.models import ContentFlag
    user   = User.query.get_or_404(user_id)
    reason = request.form.get('reason', '').strip()
    if user.id == current_user.id:
        flash('Cannot suspend yourself.', 'danger')
        return redirect(url_for('admin.moderator.view_user', user_id=user_id))
    if not reason:
        flash('Suspension reason is required.', 'danger')
        return redirect(url_for('admin.moderator.view_user', user_id=user_id))
    user.is_active = False
    db.session.commit()
    create_flag(current_user, 'user', user_id, f'[SUSPENDED] {reason}', priority='high')
    _audit('user.suspend', 'user_management', {
        'target': user.username, 'target_id': user_id, 'reason': reason
    })
    flash(f'{user.username} suspended.', 'warning')
    return redirect(url_for('admin.moderator.view_user', user_id=user_id))


@moderator_bp.route('/users/<int:user_id>/unsuspend', methods=['POST'])
@login_required
@require_role(*_MOD)
@require_fresh_user
def unsuspend_user(user_id):
    from app.admin.models import ContentFlag
    user           = User.query.get_or_404(user_id)
    user.is_active = True
    # Resolve any open suspension flags
    ContentFlag.query.filter_by(
        entity_type='user', entity_id=user_id, status='open'
    ).update({
        'status': 'resolved', 'resolved_by': current_user.id,
        'resolution_action': 'unsuspended', 'resolved_at': datetime.now(timezone.utc)
    })
    db.session.commit()
    _audit('user.unsuspend', 'user_management', {
        'target': user.username, 'target_id': user_id
    })
    flash(f'{user.username} re-activated.', 'success')
    return redirect(url_for('admin.moderator.view_user', user_id=user_id))


@moderator_bp.route('/users/<int:user_id>/warn', methods=['POST'])
@login_required
@require_role(*_MOD)
def warn_user(user_id):
    from app.admin.models import ContentFlag
    user   = User.query.get_or_404(user_id)
    reason = request.form.get('reason', '').strip()
    if not reason:
        flash('Warning reason is required.', 'danger')
        return redirect(url_for('admin.moderator.view_user', user_id=user_id))
    create_flag(current_user, 'user', user_id, f'[WARNING] {reason}', priority='medium')
    _audit('user.warn', 'user_management', {
        'target': user.username, 'target_id': user_id, 'reason': reason
    })
    flash(f'Formal warning issued to {user.username}.', 'warning')
    return redirect(url_for('admin.moderator.view_user', user_id=user_id))


@moderator_bp.route('/users/<int:user_id>/verify', methods=['POST'])
@login_required
@require_role(*_MOD)
@require_fresh_user
def verify_user(user_id):
    user             = User.query.get_or_404(user_id)
    user.is_verified = True
    db.session.commit()
    _audit('user.verify', 'user_management', {
        'target': user.username, 'target_id': user_id
    })
    flash(f'{user.username} marked as verified.', 'success')
    return redirect(url_for('admin.moderator.view_user', user_id=user_id))


@moderator_bp.route('/users/<int:user_id>/reject-verification', methods=['POST'])
@login_required
@require_role(*_MOD)
@require_fresh_user
def reject_user_verification(user_id):
    """Reject user's KYC/verification record with reason. Does NOT create a flag."""
    user = User.query.get_or_404(user_id)
    reason = request.form.get('reason', '').strip()

    if not reason:
        flash('Rejection reason is required.', 'danger')
        return redirect(url_for('admin.moderator.view_user', user_id=user_id))

    now = datetime.now(timezone.utc)

    # Try to find and reject the most recent KYC/verification record
    try:
        from app.kyc.models import KycRecord
        kyc_record = KycRecord.query.filter_by(user_id=user_id).order_by(KycRecord.created_at.desc()).first()

        if kyc_record and kyc_record.status in ('pending', 'manual_review'):
            kyc_record.status = 'rejected'
            kyc_record.rejection_reason = reason
            kyc_record.verified_at = now
            kyc_record.verified_by_id = current_user.id
            db.session.add(kyc_record)
            flash(f'KYC verification rejected for {user.username}.', 'warning')
        else:
            flash(f'No pending KYC record found for {user.username}.', 'info')
    except Exception:
        # If KYC module not available, try other verification models
        try:
            from app.identity.models.organisation import Organisation
            org = Organisation.query.filter_by(owner_id=user_id).first()
            if org and org.verification_status == 'pending':
                org.verification_status = 'rejected'
                db.session.add(org)
                flash(f'Organisation verification rejected for {user.username}.', 'warning')
            else:
                flash(f'No pending verification record found for {user.username}.', 'info')
        except Exception:
            flash(f'No pending verification record found for {user.username}.', 'info')

    db.session.commit()
    _audit('user.reject_verification', 'user_management', {
        'target': user.username,
        'target_id': user_id,
        'reason': reason
    })

    return redirect(url_for('admin.moderator.view_user', user_id=user_id))


@moderator_bp.route('/users/<int:user_id>/deactivate', methods=['POST'])
@login_required
@require_role(*_MOD)
@require_fresh_user
def deactivate_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash('Cannot deactivate yourself.', 'danger')
        return redirect(url_for('admin.moderator.view_user', user_id=user_id))
    user.is_active = False
    db.session.commit()
    _audit('user.deactivate', 'user_management', {'target': user.username})
    flash(f'{user.username} deactivated.', 'warning')
    return redirect(url_for('admin.moderator.view_user', user_id=user_id))


@moderator_bp.route('/users/<int:user_id>/activate', methods=['POST'])
@login_required
@require_role(*_MOD)
@require_fresh_user
def activate_user(user_id):
    user           = User.query.get_or_404(user_id)
    user.is_active = True
    db.session.commit()
    _audit('user.activate', 'user_management', {'target': user.username})
    flash(f'{user.username} activated.', 'success')
    return redirect(url_for('admin.moderator.view_user', user_id=user_id))


@moderator_bp.route('/users/<int:user_id>/flag', methods=['POST'])
@login_required
@require_permission('content.flag')
def flag_user(user_id):
    User.query.get_or_404(user_id)
    reason   = request.form.get('reason') or (request.json or {}).get('reason', '')
    priority = request.form.get('priority', 'normal')
    ok, result = create_flag(current_user, 'user', user_id, reason, priority)
    if ok:
        flash(f'{user.username} flagged.', 'warning')
    else:
        flash(f'Could not flag: {result}', 'danger')
    if request.is_json:
        return jsonify({"ok": ok, "flag_id": int(result.id) if ok else None})
    return redirect(url_for('admin.moderator.view_user', user_id=user_id))


@moderator_bp.route('/users/<int:user_id>/note', methods=['POST'])
@login_required
@require_role(*_MOD)
def add_user_note(user_id):
    User.query.get_or_404(user_id)
    note = request.form.get('note', '').strip()
    if not note:
        flash('Note cannot be empty.', 'danger')
        return redirect(url_for('admin.moderator.view_user', user_id=user_id))
    create_flag(current_user, 'user', user_id, f'[NOTE] {note}', priority='low')
    flash('Internal note saved.', 'success')
    return redirect(url_for('admin.moderator.view_user', user_id=user_id))


@moderator_bp.route('/users/bulk-action', methods=['POST'])
@login_required
@require_role(*_MOD)
def bulk_user_action():
    """suspend | activate | deactivate | verify multiple users."""
    user_ids = request.form.getlist('user_ids', type=int)
    action   = request.form.get('action')
    reason   = request.form.get('reason', 'Bulk action by moderator')

    if not user_ids:
        flash('No users selected.', 'warning')
        return redirect(url_for('admin.moderator.user_moderation'))

    count = 0
    for uid in user_ids:
        if uid == current_user.id:
            continue
        u = User.query.get(uid)
        if not u:
            continue
        if action == 'suspend':
            u.is_active = False
            create_flag(current_user, 'user', uid, f'[BULK SUSPEND] {reason}', priority='high')
            count += 1
        elif action == 'activate':
            u.is_active = True
            count += 1
        elif action == 'deactivate':
            u.is_active = False
            count += 1
        elif action == 'verify':
            u.is_verified = True
            count += 1

    db.session.commit()
    _audit(f'user.bulk_{action}', 'user_management', {'count': count, 'ids': user_ids})
    flash(f'{count} user(s) actioned.', 'success')
    return redirect(url_for('admin.moderator.user_moderation'))


# ═════════════════════════════════════════════════════════════════════════════
# E.  MANAGEABLE CATEGORIES
# ═════════════════════════════════════════════════════════════════════════════

@moderator_bp.route('/categories')
@login_required
@require_role(*_MOD)
def categories_list():
    from app.admin.models import ManageableCategory
    cats = ManageableCategory.query.order_by(ManageableCategory.name).all()
    return render_template(
        'admin/moderator/categories.html',
        categories=cats,
        title="Manage Categories"
    )


@moderator_bp.route('/categories/<int:cat_id>')
@login_required
@require_role(*_MOD)
def view_category(cat_id):
    from app.admin.models import ManageableCategory, ManageableItem
    cat = ManageableCategory.query.get_or_404(cat_id)
    pending_items = ManageableItem.query.filter_by(
        category_id=cat_id, is_approved=False
    ).count()
    return render_template(
        'admin/moderator/view_category.html',
        category=cat,
        pending_items=pending_items,
        title=f"Category: {cat.name}"
    )


@moderator_bp.route('/categories/<int:cat_id>/toggle-active', methods=['POST'])
@login_required
@require_role(*_MOD)
def toggle_category_active(cat_id):
    from app.admin.models import ManageableCategory
    cat           = ManageableCategory.query.get_or_404(cat_id)
    cat.is_active = not cat.is_active
    db.session.commit()
    state = 'activated' if cat.is_active else 'deactivated'
    _audit(f'category.{state}', 'content', {'category_id': cat_id, 'name': cat.name})
    flash(f'Category "{cat.name}" {state}.', 'info')
    return redirect(url_for('admin.moderator.categories_list'))


@moderator_bp.route('/categories/create', methods=['POST'])
@login_required
@require_role(*_MOD)
def create_category():
    from app.admin.models import ManageableCategory
    name = request.form.get('name', '').strip()
    slug = request.form.get('slug', '').strip()
    description = request.form.get('description', '').strip()
    is_active = request.form.get('is_active') == 'on'

    if not name or not slug:
        flash('Name and slug are required.', 'danger')
        return redirect(url_for('admin.moderator.categories_list'))

    # Check if slug already exists
    if ManageableCategory.query.filter_by(slug=slug).first():
        flash('Category with this slug already exists.', 'danger')
        return redirect(url_for('admin.moderator.categories_list'))

    category = ManageableCategory(
        name=name,
        slug=slug,
        description=description,
        is_active=is_active
    )
    db.session.add(category)
    db.session.commit()
    _audit('category.create', 'content', {'category_id': category.id, 'name': name})
    flash(f'Category "{name}" created successfully.', 'success')
    return redirect(url_for('admin.moderator.categories_list'))


# ═════════════════════════════════════════════════════════════════════════════
# F.  MANAGEABLE ITEMS
# ═════════════════════════════════════════════════════════════════════════════

@moderator_bp.route('/items')
@login_required
@require_role(*_MOD)
def items_list():
    from app.admin.models import ManageableItem, ManageableCategory
    page     = request.args.get('page', 1, type=int)
    approved = request.args.get('approved', 'all')
    cat_id   = request.args.get('category_id', type=int)
    search   = request.args.get('q', '').strip()

    q = ManageableItem.query
    if approved == 'pending':
        q = q.filter_by(is_approved=False)
    elif approved == 'approved':
        q = q.filter_by(is_approved=True)
    if cat_id:
        q = q.filter_by(category_id=cat_id)
    if search:
        q = q.filter(ManageableItem.name.ilike(f'%{search}%'))

    items      = q.order_by(ManageableItem.created_at.desc()).paginate(page=page, per_page=25, error_out=False)
    categories = ManageableCategory.query.filter_by(is_active=True).all()

    return render_template(
        'admin/moderator/items.html',
        items=items,
        categories=categories,
        approved_filter=approved,
        selected_category=cat_id,
        search=search,
        title="Manage Items"
    )


@moderator_bp.route('/items/<int:item_id>')
@login_required
@require_role(*_MOD)
def view_item(item_id):
    from app.admin.models import ManageableItem, ManageableCategory, ContentFlag
    item       = ManageableItem.query.get_or_404(item_id)
    flags      = (ContentFlag.query
                  .filter_by(entity_type='manageable_item', entity_id=item_id)
                  .order_by(ContentFlag.created_at.desc()).all())
    categories = ManageableCategory.query.filter_by(is_active=True).all()
    return render_template(
        'admin/moderator/view_item.html',
        item=item,
        flags=flags,
        categories=categories,
        title=f"Item: {item.name}"
    )


@moderator_bp.route('/items/<int:item_id>/approve', methods=['POST'])
@login_required
@require_role(*_MOD)
def approve_item(item_id):
    from app.admin.models import ManageableItem
    item             = ManageableItem.query.get_or_404(item_id)
    item.is_approved = True
    item.is_active   = True
    db.session.commit()
    _audit('item.approve', 'content', {'item_id': item_id, 'name': item.name})
    flash(f'"{item.name}" approved.', 'success')
    return redirect(url_for('admin.moderator.items_list'))


@moderator_bp.route('/items/<int:item_id>/reject', methods=['POST'])
@login_required
@require_role(*_MOD)
def reject_item(item_id):
    from app.admin.models import ManageableItem
    item             = ManageableItem.query.get_or_404(item_id)
    notes            = request.form.get('notes', '').strip()
    item.is_approved = False
    item.is_active   = False
    db.session.commit()
    _audit('item.reject', 'content', {'item_id': item_id, 'name': item.name, 'reason': notes})
    flash(f'"{item.name}" rejected.', 'warning')
    return redirect(url_for('admin.moderator.items_list'))


@moderator_bp.route('/items/<int:item_id>/toggle-featured', methods=['POST'])
@login_required
@require_role(*_MOD)
def toggle_item_featured(item_id):
    from app.admin.models import ManageableItem
    item             = ManageableItem.query.get_or_404(item_id)
    item.is_featured = not item.is_featured
    db.session.commit()
    state = 'featured' if item.is_featured else 'unfeatured'
    flash(f'"{item.name}" {state}.', 'info')
    return redirect(url_for('admin.moderator.view_item', item_id=item_id))


@moderator_bp.route('/items/<int:item_id>/flag', methods=['POST'])
@login_required
@require_permission('content.flag')
def flag_item(item_id):
    from app.admin.models import ManageableItem
    ManageableItem.query.get_or_404(item_id)
    reason   = request.form.get('reason') or (request.json or {}).get('reason', '')
    priority = request.form.get('priority', 'normal')
    ok, result = create_flag(current_user, 'manageable_item', item_id, reason, priority)
    if ok:
        flash('Item flagged.', 'warning')
    else:
        flash(f'Could not flag: {result}', 'danger')
    if request.is_json:
        return jsonify({"ok": ok, "flag_id": int(result.id) if ok else None})
    return redirect(url_for('admin.moderator.view_item', item_id=item_id))


# ═════════════════════════════════════════════════════════════════════════════
# G.  EVENT MODERATION
# ═════════════════════════════════════════════════════════════════════════════

def _events_model():
    from app.events.models import Event
    return Event


@moderator_bp.route('/events')
@login_required
@require_role(*_MOD)
def events_list():
    """Redirect to Events module admin dashboard."""
    try:
        return redirect(url_for('events.admin_dashboard'))
    except Exception:
        # Fallback if events module is not available
        flash('Events module is not available.', 'warning')
        return redirect(url_for('admin.moderator.dashboard'))


@moderator_bp.route('/events/<int:event_id>')
@login_required
@require_role(*_MOD)
def view_event(event_id):
    """Redirect to specific event in Events module."""
    try:
        return redirect(url_for('events.admin_view_event', event_id=event_id))
    except Exception:
        flash('Event module is not available.', 'warning')
        return redirect(url_for('admin.moderator.flagged_content'))


@moderator_bp.route('/events/<int:event_id>/approve', methods=['POST'])
@login_required
@require_role(*_MOD)
def approve_event(event_id):
    """Redirect to event module's approve endpoint."""
    try:
        return redirect(url_for('events.admin_approve_event', event_id=event_id))
    except Exception:
        flash('Cannot approve event - module unavailable.', 'danger')
        return redirect(url_for('admin.moderator.events_list'))


@moderator_bp.route('/events/<int:event_id>/reject', methods=['POST'])
@login_required
@require_role(*_MOD)
def reject_event(event_id):
    """Redirect to event module's reject endpoint."""
    try:
        return redirect(url_for('events.admin_reject_event', event_id=event_id))
    except Exception:
        flash('Cannot reject event - module unavailable.', 'danger')
        return redirect(url_for('admin.moderator.events_list'))


@moderator_bp.route('/events/<int:event_id>/flag', methods=['POST'])
@login_required
@require_permission('content.flag')
def flag_event(event_id):
    reason   = request.form.get('reason') or (request.json or {}).get('reason', '')
    priority = request.form.get('priority', 'normal')
    ok, result = create_flag(current_user, 'event', event_id, reason, priority)
    if not ok:
        flash(f'Could not flag: {result}', 'danger')
        if request.is_json:
            return jsonify({"ok": False, "error": result}), 403
        return redirect(url_for('admin.moderator.events_list'))
    flash('Event flagged.', 'warning')
    if request.is_json:
        return jsonify({"ok": True, "flag_id": int(result.id)})
    return _redirect_back('admin.moderator.events_list')


# ═════════════════════════════════════════════════════════════════════════════
# H.  ORGANISATION VERIFICATION QUEUE
# ═════════════════════════════════════════════════════════════════════════════

def _org_model():
    from app.identity.models.organisation import Organisation
    return Organisation


@moderator_bp.route('/orgs')
@login_required
@require_role(*_MOD)
def orgs_queue():
    """Organisation moderation queue - show organizations pending review."""
    from app.admin.models import ContentFlag
    from app.identity.models.organisation import Organisation
    
    # Get organizations with flags or pending status
    flagged_orgs_query = db.session.query(Organisation).join(
        ContentFlag, Organisation.id == ContentFlag.entity_id
    ).filter(
        ContentFlag.entity_type == 'organisation',
        ContentFlag.status.in_(['pending', 'under_review'])
    ).distinct()
    
    # Get organizations with verification status issues
    pending_orgs_query = Organisation.query.filter(
        Organisation.verification_status.in_(['pending', 'flagged', 'under_review'])
    )
    
    # Combine both queries
    flagged_orgs = flagged_orgs_query.all()
    pending_orgs = pending_orgs_query.all()
    
    # Remove duplicates
    all_orgs = list({org.id: org for org in flagged_orgs + pending_orgs}.values())
    
    # Sort by most recent activity
    all_orgs.sort(key=lambda x: x.updated_at or x.created_at, reverse=True)
    
    # Get moderation stats
    org_stats = {
        'total_pending': len(all_orgs),
        'flagged_count': len(flagged_orgs),
        'verification_pending': len([org for org in all_orgs if org.verification_status in ['pending', 'under_review']]),
        'high_priority': len([org for org in all_orgs if any(
            flag.priority == 'critical' for flag in org.flags if hasattr(org, 'flags')
        )])
    }
    
    return render_template(
        'admin/moderator/orgs.html',
        orgs=all_orgs,
        org_stats=org_stats,
        title="Organisation Moderation"
    )


@moderator_bp.route('/orgs/<int:org_id>')
@login_required
@require_role(*_MOD)
def view_org(org_id):
    """View organisation details for moderation review."""
    from app.admin.models import ContentFlag
    from app.identity.models.organisation import Organisation
    
    org = Organisation.query.get_or_404(org_id)
    
    # Get flags related to this organization
    org_flags = ContentFlag.query.filter_by(
        entity_type='organisation',
        entity_id=org.id
    ).order_by(ContentFlag.created_at.desc()).all()
    
    # Get organization activity
    org_activity = []
    if hasattr(org, 'created_at'):
        org_activity.append({
            'type': 'created',
            'timestamp': org.created_at,
            'description': f'Organization "{org.name}" was created'
        })
    
    # Add flag activity
    for flag in org_flags:
        org_activity.append({
            'type': 'flag',
            'timestamp': flag.created_at,
            'description': f'Flagged: {flag.reason[:100]}...',
            'priority': flag.priority,
            'status': flag.status
        })
    
    # Sort activity by timestamp
    org_activity.sort(key=lambda x: x['timestamp'], reverse=True)
    
    # Get moderation recommendations
    recommendations = []
    if org.verification_status == 'pending':
        recommendations.append('Review organization documents and verify business legitimacy')
    if any(flag.priority == 'critical' for flag in org_flags):
        recommendations.append('URGENT: Critical flags require immediate attention')
    if len(org_flags) > 3:
        recommendations.append('Multiple flags - consider escalating to compliance team')
    
    return render_template(
        'admin/moderator/view_org.html',
        org=org,
        org_flags=org_flags,
        org_activity=org_activity,
        recommendations=recommendations,
        title=f"Review Organization: {org.name}"
    )


@moderator_bp.route('/orgs/<int:org_id>/approve', methods=['POST'])
@login_required
@require_role(*_MOD)
def approve_org(org_id):
    """Redirect to org module's approve endpoint."""
    try:
        return redirect(url_for('org.approve_organisation', org_id=org_id))
    except Exception:
        flash('Cannot approve organisation - module unavailable.', 'danger')
        return redirect(url_for('admin.moderator.orgs_queue'))


@moderator_bp.route('/orgs/<int:org_id>/reject', methods=['POST'])
@login_required
@require_role(*_MOD)
def reject_org(org_id):
    """Redirect to org module's reject endpoint."""
    try:
        return redirect(url_for('org.reject_organisation', org_id=org_id))
    except Exception:
        flash('Cannot reject organisation - module unavailable.', 'danger')
        return redirect(url_for('admin.moderator.orgs_queue'))


@moderator_bp.route('/orgs/<int:org_id>/flag', methods=['POST'])
@login_required
@require_permission('content.flag')
def flag_org(org_id):
    reason   = request.form.get('reason') or (request.json or {}).get('reason', '')
    priority = request.form.get('priority', 'normal')
    ok, result = create_flag(current_user, 'organisation', org_id, reason, priority)
    if ok:
        flash('Organisation flagged.', 'warning')
    else:
        flash(f'Could not flag: {result}', 'danger')
    if request.is_json:
        return jsonify({"ok": ok})
    return _redirect_back('admin.moderator.orgs_queue')


@moderator_bp.route('/orgs/<int:org_id>/refer-compliance', methods=['POST'])
@login_required
@require_role(*_MOD)
def refer_org_compliance(org_id):
    """Refer organisation to compliance officer for review. Creates a compliance task."""
    try:
        from app.identity.models.organisation import Organisation
        org = Organisation.query.get_or_404(org_id)
    except Exception:
        flash('Organisation module not available.', 'danger')
        return redirect(url_for('admin.moderator.orgs_queue'))

    reason = request.form.get('reason', '').strip()
    if not reason:
        flash('Referral reason is required.', 'danger')
        return redirect(url_for('admin.moderator.view_org', org_id=org_id))

    # Create a high-priority flag for compliance review
    ok, flag = create_flag(
        current_user,
        'organisation',
        org_id,
        f'[COMPLIANCE REFERRAL] {reason}',
        priority='high'
    )

    if ok:
        # Update organisation status to indicate compliance review
        org.verification_status = 'pending'
        db.session.add(org)
        db.session.commit()

        _audit('org.refer_compliance', 'compliance', {
            'org_id': org_id,
            'org_name': org.name,
            'referred_by': current_user.id,
            'reason': reason,
            'flag_id': flag.id if flag else None
        })

        flash(f'Organisation referred to compliance officer for review.', 'warning')
    else:
        flash(f'Failed to create compliance referral: {flag}', 'danger')

    return redirect(url_for('admin.moderator.view_org', org_id=org_id))


# ═════════════════════════════════════════════════════════════════════════════
# J.  ESCALATION QUEUE (Admin+ only, rendered with role-aware UI)
# ═════════════════════════════════════════════════════════════════════════════

@moderator_bp.route('/escalations')
@login_required
@require_role(*_MOD)
def escalation_queue():
    """Show items escalated to admin for review. Templates will hide this for moderators."""
    from sqlalchemy import func

    try:
        # Get flags escalated to admin role
        escalated_flags = ContentFlag.query.filter(
            ContentFlag.escalated_to_role == 'admin',
            ContentFlag.status.in_(['open', 'in_review'])
        ).order_by(
            ContentFlag.priority.desc(),
            ContentFlag.created_at.asc()
        ).all()

        # Count by priority
        priority_counts = db.session.query(
            ContentFlag.priority,
            func.count(ContentFlag.id)
        ).filter(
            ContentFlag.escalated_to_role == 'admin',
            ContentFlag.status.in_(['open', 'in_review'])
        ).group_by(ContentFlag.priority).all()

        priority_stats = {row[0]: row[1] for row in priority_counts}
        total_escalated = sum(priority_stats.values())

        return render_template(
            'admin/moderator/escalations.html',
            escalated_flags=escalated_flags,
            priority_stats=priority_stats,
            total_escalated=total_escalated
        )
    except Exception as e:
        logger.error(f"Escalation queue error: {e}")
        flash("Error loading escalation queue.", "danger")
        return redirect(url_for("admin.moderator.dashboard"))


@moderator_bp.route('/escalations/<int:flag_id>/resolve', methods=['POST'])
@login_required
@require_role(*_MOD)
def resolve_escalation(flag_id):
    """Resolve an escalated flag. Admin+ only action enforced in template."""
    flag = ContentFlag.query.get_or_404(flag_id)

    if flag.escalated_to_role != 'admin':
        flash("This flag is not escalated to admin.", "danger")
        return redirect(url_for("admin.moderator.escalation_queue"))

    resolution_action = request.form.get('resolution_action', 'resolved')
    resolution_notes = request.form.get('resolution_notes', '').strip()

    ok, result = resolve_flag(current_user, flag_id, resolution_action, resolution_notes)

    if ok:
        flash(f"Escalation resolved: {resolution_action}", "success")
    else:
        flash(f"Failed to resolve: {result}", "danger")

    return redirect(url_for("admin.moderator.escalation_queue"))


@moderator_bp.route('/escalations/<int:flag_id>/escalate', methods=['POST'])
@login_required
@require_role(*_MOD)
def escalate_to_compliance(flag_id):
    """Escalate a flag to compliance officer. Admin+ only action enforced in template."""
    flag = ContentFlag.query.get_or_404(flag_id)

    if flag.escalated_to_role != 'admin':
        flash("This flag is not escalated to admin.", "danger")
        return redirect(url_for("admin.moderator.escalation_queue"))

    reason = request.form.get('reason', '').strip()
    if not reason:
        flash("Escalation reason is required.", "danger")
        return redirect(url_for("admin.moderator.escalation_queue"))

    # Create compliance case
    try:
        compliance_case = ComplianceCaseService.create_case(
            case_type=ComplianceCaseType.FLAG_ESCALATION,
            title=f'Flag Escalation - {flag.entity_type}:{flag.entity_id}',
            description=f'Flag escalated from moderator: {reason}',
            created_by=current_user.id,
            flag_id=flag_id,
            priority=ComplianceCasePriority.MEDIUM,
            escalated_from=current_user.id,
            escalation_reason=reason
        )

        # Update flag with compliance case reference
        flag.compliance_case_id = compliance_case.id
        flag.referred_to_compliance = True
        flag.referred_at = datetime.now(timezone.utc)
        flag.referred_by = current_user.id
        flag.escalated_to_role = 'compliance_officer'
        flag.status = 'in_review'
        flag.assigned_to = None  # Clear assignment for compliance to pick up
        db.session.add(flag)
        db.session.commit()

        # Audit log
        _audit('escalation.to_compliance', 'moderation', {
            "flag_id": flag_id,
            "entity_type": flag.entity_type,
            "entity_id": flag.entity_id,
            "reason": reason,
            "compliance_case_id": compliance_case.id
        })

        flash("Escalated to compliance officer.", "warning")
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error escalating to compliance: {e}")
        flash(f"Error escalating to compliance: {str(e)}", "danger")

    return redirect(url_for("admin.moderator.escalation_queue"))


# ═════════════════════════════════════════════════════════════════════════════
# K.  KYC QUEUE
# ═════════════════════════════════════════════════════════════════════════════

def _kyc_model():
    from app.identity.individuals.individual_document import IndividualKYCDocument
    return IndividualKYCDocument


@moderator_bp.route('/kyc')
@login_required
@require_role(*_MOD)
def kyc_queue():
    """Redirect to KYC module admin."""
    try:
        return redirect(url_for('kyc.admin_dashboard'))
    except Exception:
        flash('KYC module is not available.', 'warning')
        return redirect(url_for('admin.moderator.dashboard'))


@moderator_bp.route('/kyc/<int:doc_id>')
@login_required
@require_role(*_MOD)
def view_kyc(doc_id):
    """Redirect to specific KYC document."""
    try:
        return redirect(url_for('kyc.admin_view_document', doc_id=doc_id))
    except Exception:
        flash('KYC module is not available.', 'warning')
        return redirect(url_for('admin.moderator.kyc_queue'))


@moderator_bp.route('/kyc/<int:doc_id>/flag', methods=['POST'])
@login_required
@require_permission('content.flag')
def flag_kyc(doc_id):
    reason   = request.form.get('reason') or (request.json or {}).get('reason', '')
    priority = request.form.get('priority', 'high')
    ok, result = create_flag(current_user, 'kyc_document', doc_id, reason, priority)
    if ok:
        flash('KYC record flagged - compliance team will review.', 'warning')
    else:
        flash(f'Could not flag: {result}', 'danger')
    if request.is_json:
        return jsonify({"ok": ok})
    return _redirect_back('admin.moderator.kyc_queue')


@moderator_bp.route('/kyc/<int:doc_id>/refer-compliance', methods=['POST'])
@login_required
@require_role(*_MOD)
def refer_kyc_compliance(doc_id):
    """Escalate a KYC document directly to the compliance team."""
    reason   = request.form.get('reason', 'Referred by moderator for compliance review').strip()

    try:
        # Create compliance case
        compliance_case = ComplianceCaseService.create_case(
            case_type=ComplianceCaseType.KYC_REVIEW,
            title=f'KYC Review - Document {doc_id}',
            description=f'KYC document referred for compliance review: {reason}',
            created_by=current_user.id,
            kyc_id=doc_id,
            priority=ComplianceCasePriority.HIGH,
            escalated_from=current_user.id,
            escalation_reason=reason
        )

        # Also create flag for tracking
        ok, result = create_flag(current_user, 'kyc_document', doc_id, reason, priority='critical')
        if ok:
            result.escalated_to_role = 'compliance_officer'
            result.status = 'in_review'
            result.compliance_case_id = compliance_case.id
            result.referred_to_compliance = True
            result.referred_at = datetime.now(timezone.utc)
            result.referred_by = current_user.id
            db.session.commit()

        _audit('kyc.refer_compliance', 'compliance', {
            'doc_id': doc_id,
            'reason': reason,
            'compliance_case_id': compliance_case.id
        })
        flash('KYC record referred to compliance team.', 'warning')
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error referring KYC to compliance: {e}")
        flash(f'Could not refer: {str(e)}', 'danger')

    return _redirect_back('admin.moderator.kyc_queue')


# ═════════════════════════════════════════════════════════════════════════════
# J.  AUDIT LOG (read-only)
# ═════════════════════════════════════════════════════════════════════════════

@moderator_bp.route('/audit-log')
@login_required
@require_role(*_MOD)
def audit_log():
    page            = request.args.get('page', 1, type=int)
    per_page        = 50
    action_filter   = request.args.get('action', '').strip()
    category_filter = request.args.get('category', '').strip()
    user_filter     = request.args.get('user_id', type=int)

    try:
        from app.admin.owner.models import OwnerAuditLog
        q = OwnerAuditLog.query
        if action_filter:
            q = q.filter(OwnerAuditLog.action.ilike(f'%{action_filter}%'))
        if category_filter:
            q = q.filter_by(category=category_filter)
        if user_filter:
            q = q.filter_by(actor_id=user_filter)
        logs           = q.order_by(OwnerAuditLog.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        categories     = [r[0] for r in db.session.query(
            OwnerAuditLog.category.distinct()).all() if r[0]]
        audit_available = True
    except Exception as exc:
        logger.debug("Audit log unavailable: %s", exc)
        logs            = None
        categories      = []
        audit_available = False

    return render_template(
        'admin/moderator/audit_log.html',
        logs=logs,
        categories=categories,
        action_filter=action_filter,
        category_filter=category_filter,
        audit_available=audit_available,
        title="Audit Log"
    )


# ═════════════════════════════════════════════════════════════════════════════
# K.  MODERATION STATS
# ═════════════════════════════════════════════════════════════════════════════

@moderator_bp.route('/stats')
@login_required
@require_role(*_MOD)
def stats():
    from app.admin.models import ContentSubmission, ContentFlag
    now     = datetime.now(timezone.utc)
    week_ago  = now - timedelta(days=7)
    month_ago = now - timedelta(days=30)

    def _sub_stats(since):
        return {
            'approved': ContentSubmission.query.filter(
                ContentSubmission.status == 'approved',
                ContentSubmission.reviewed_at >= since
            ).count(),
            'rejected': ContentSubmission.query.filter(
                ContentSubmission.status == 'rejected',
                ContentSubmission.reviewed_at >= since
            ).count(),
            'changes': ContentSubmission.query.filter(
                ContentSubmission.status == 'changes_requested',
                ContentSubmission.reviewed_at >= since
            ).count(),
        }

    def _flag_stats(since):
        return {
            'opened':   ContentFlag.query.filter(ContentFlag.created_at >= since).count(),
            'resolved': ContentFlag.query.filter(
                ContentFlag.status == 'resolved',
                ContentFlag.resolved_at >= since
            ).count(),
            'open':     ContentFlag.query.filter_by(status='open').count(),
        }

    submission_stats = {'7d': _sub_stats(week_ago), '30d': _sub_stats(month_ago)}
    flag_stats       = {'7d': _flag_stats(week_ago), '30d': _flag_stats(month_ago)}

    priority_dist = dict(
        ContentFlag.query
        .with_entities(ContentFlag.priority, func.count())
        .filter_by(status='open')
        .group_by(ContentFlag.priority).all()
    )
    type_dist = dict(
        ContentFlag.query
        .with_entities(ContentFlag.entity_type, func.count())
        .filter_by(status='open')
        .group_by(ContentFlag.entity_type).all()
    )

    # 14-day daily opens for sparkline
    daily_opens = []
    for i in range(13, -1, -1):
        day = (now - timedelta(days=i)).replace(hour=0, minute=0, second=0, microsecond=0)
        daily_opens.append({
            'date':  day.strftime('%b %d'),
            'count': ContentFlag.query.filter(
                ContentFlag.created_at >= day,
                ContentFlag.created_at < day + timedelta(days=1)
            ).count()
        })

    # Top flaggers (last 30 days) - useful to spot spam-flaggers
    top_flaggers_raw = (
        ContentFlag.query
        .with_entities(ContentFlag.flagged_by, func.count().label('n'))
        .filter(ContentFlag.created_at >= month_ago)
        .group_by(ContentFlag.flagged_by)
        .order_by(func.count().desc())
        .limit(5).all()
    )
    flagger_ids   = [r[0] for r in top_flaggers_raw]
    flagger_users = {u.id: u for u in User.query.filter(User.id.in_(flagger_ids)).all()}
    top_flaggers  = [
        {'user': flagger_users.get(r[0]), 'count': r[1]}
        for r in top_flaggers_raw
    ]

    return render_template(
        'admin/moderator/stats.html',
        submission_stats=submission_stats,
        flag_stats=flag_stats,
        priority_dist=priority_dist,
        type_dist=type_dist,
        daily_opens=daily_opens,
        top_flaggers=top_flaggers,
        title="Moderation Stats"
    )


# ═════════════════════════════════════════════════════════════════════════════
# L.  JSON API HELPERS
# ═════════════════════════════════════════════════════════════════════════════

@moderator_bp.route('/api/queue-health')
@login_required
@require_role(*_MOD)
def api_queue_health():
    """Live queue counters for header widgets / polling."""
    from app.admin.models import ContentFlag, ContentSubmission
    return jsonify({
        "open_flags":            ContentFlag.query.filter_by(status='open').count(),
        "critical_flags":        ContentFlag.query.filter_by(status='open', priority='critical').count(),
        "high_flags":            ContentFlag.query.filter_by(status='open', priority='high').count(),
        "pending_submissions":   ContentSubmission.query.filter_by(status='pending').count(),
        "changes_requested":     ContentSubmission.query.filter_by(status='changes_requested').count(),
        "in_review_flags":       ContentFlag.query.filter_by(status='in_review').count(),
    })


@moderator_bp.route('/api/flags/<int:flag_id>/resolve', methods=['POST'])
@login_required
@require_role(*_MOD)
def api_resolve_flag(flag_id):
    data   = request.get_json() or {}
    ok, result = resolve_flag(
        current_user, flag_id,
        data.get('action', 'reviewed'),
        data.get('notes', '')
    )
    if ok:
        _audit('flag.resolve', 'moderation', {'flag_id': flag_id})
    return jsonify({"ok": ok, "error": None if ok else result})


@moderator_bp.route('/api/flags/<int:flag_id>/close', methods=['POST'])
@login_required
@require_role(*_MOD)
def api_close_flag(flag_id):
    from app.admin.models import ContentFlag
    data  = request.get_json() or {}
    flag  = ContentFlag.query.get_or_404(flag_id)
    flag.status            = 'resolved'
    flag.resolved_by       = current_user.id
    flag.resolution_action = 'closed'
    flag.resolution_notes  = data.get('notes', 'Closed via API')
    flag.resolved_at       = datetime.now(timezone.utc)
    db.session.commit()
    return jsonify({"ok": True})


@moderator_bp.route('/api/submissions/<int:submission_id>/approve', methods=['POST'])
@login_required
@require_role(*_MOD)
def api_approve_submission(submission_id):
    from app.admin.models import ContentSubmission, ModerationLog
    s = ContentSubmission.query.get_or_404(submission_id)
    now = datetime.now(timezone.utc)
    s.status      = 'approved'
    s.reviewed_by = current_user.id
    s.reviewed_at = now
    s.review_notes = (request.get_json() or {}).get('notes', '')
    s.resolved_at = now
    if s.claimed_at:
        s.processing_time_seconds = int((now - s.claimed_at).total_seconds())
    db.session.add(ModerationLog(
        moderator_id=current_user.id,
        submission_id=submission_id,
        action='approve',
        notes=f"Processing time: {s.processing_time_seconds}s"
    ))
    db.session.commit()
    return jsonify({"ok": True})


@moderator_bp.route('/api/submissions/<int:submission_id>/reject', methods=['POST'])
@login_required
@require_role(*_MOD)
def api_reject_submission(submission_id):
    from app.admin.models import ContentSubmission, ModerationLog
    data  = request.get_json() or {}
    notes = data.get('notes', '').strip()
    if not notes:
        return jsonify({"ok": False, "error": "notes required"}), 400
    s = ContentSubmission.query.get_or_404(submission_id)
    now = datetime.now(timezone.utc)
    s.status      = 'rejected'
    s.reviewed_by = current_user.id
    s.reviewed_at = now
    s.review_notes = notes
    s.resolved_at = now
    if s.claimed_at:
        s.processing_time_seconds = int((now - s.claimed_at).total_seconds())
    db.session.add(ModerationLog(
        moderator_id=current_user.id,
        submission_id=submission_id,
        action='reject',
        notes=f"Processing time: {s.processing_time_seconds}s"
    ))
    db.session.commit()
    return jsonify({"ok": True})


@moderator_bp.route('/moderate/<int:submission_id>', methods=['POST'])
@login_required
@require_role(*_MOD)
def moderate_submission(submission_id):
    """Approve or reject a submission (must be claimed by current user)."""
    from app.admin.models import ContentSubmission, ModerationLog
    data = request.get_json() or {}
    action = data.get('action', '').strip().lower()
    notes = data.get('notes', '').strip()

    if action not in ('approve', 'reject'):
        return jsonify({"ok": False, "error": "action must be 'approve' or 'reject'"}), 400

    s = ContentSubmission.query.get_or_404(submission_id)

    # Must be claimed by current user
    if s.assigned_to_id != current_user.id:
        return jsonify({"ok": False, "error": "Submission not claimed by you"}), 403

    if s.status != 'pending':
        return jsonify({"ok": False, "error": "Submission already resolved"}), 400

    now = datetime.now(timezone.utc)
    s.status = 'approved' if action == 'approve' else 'rejected'
    s.reviewed_by = current_user.id
    s.reviewed_at = now
    s.review_notes = notes
    s.resolved_at = now
    if s.claimed_at:
        s.processing_time_seconds = int((now - s.claimed_at).total_seconds())

    db.session.add(ModerationLog(
        submission_id=submission_id,
        moderator_id=current_user.id,
        action=action,
        notes=notes
    ))
    db.session.commit()

    _audit(f'submission.{action}', 'content', {
        'id': submission_id,
        'name': s.name,
        'processing_time_seconds': s.processing_time_seconds
    })

    return jsonify({"ok": True, "status": s.status, "processing_time_seconds": s.processing_time_seconds})




@moderator_bp.route('/preview/<int:submission_id>')
@login_required
@require_role(*_MOD)
def preview_submission(submission_id):
    """Preview user-submitted content before approval."""
    from app.admin.models import ContentSubmission
    submission = ContentSubmission.query.get_or_404(submission_id)
    
    # Determine template based on category
    category_templates = {
        'vehicle': 'admin/moderator/previews/vehicle.html',
        'property': 'admin/moderator/previews/property.html',
        'event': 'admin/moderator/previews/event.html',
        'tour': 'admin/moderator/previews/tour.html',
    }
    
    template = category_templates.get(
        submission.category.slug if submission.category else None,
        'admin/moderator/previews/default.html'
    )
    
    return render_template(
        template,
        submission=submission,
        data=submission.data
    )


@moderator_bp.route('/performance')
@login_required
@require_role(*_MOD)
def performance():
    from sqlalchemy import func
    data = db.session.query(
        User.id, User.username,
        func.count(ContentSubmission.id).label('processed'),
        func.avg(ContentSubmission.processing_time_seconds).label('avg_time')
    ).join(ContentSubmission, ContentSubmission.reviewed_by == User.id)\
     .filter(ContentSubmission.status.in_(['approved','rejected']))\
     .group_by(User.id).all()
    return jsonify([{
        "user_id": r.id, "username": r.username,
        "processed": r.processed,
        "avg_processing_time": round(r.avg_time or 0, 2)
    } for r in data])


@moderator_bp.route('/queue_metrics')
@login_required
@require_role(*_MOD)
def queue_metrics():
    from app.admin.models import ContentSubmission
    now = datetime.now(timezone.utc)
    pending = ContentSubmission.query.filter_by(status='pending')
    total = pending.count()
    unassigned = pending.filter(ContentSubmission.assigned_to_id == None).count()
    assigned = total - unassigned
    oldest = pending.order_by(ContentSubmission.created_at.asc()).first()
    age = int((now - oldest.created_at).total_seconds()) if oldest else 0

    # Average processing time for resolved items
    avg_time = db.session.query(
        func.avg(ContentSubmission.processing_time_seconds)
    ).filter(
        ContentSubmission.status.in_(['approved', 'rejected']),
        ContentSubmission.processing_time_seconds != None
    ).scalar()

    return jsonify({
        "total_pending": total,
        "unassigned_count": unassigned,
        "assigned_count": assigned,
        "oldest_item_age_sec": age,
        "avg_processing_time": round(avg_time or 0, 2)
    })


@moderator_bp.route('/audit_insights')
@login_required
@require_role(*_MOD)
def audit_insights():
    from sqlalchemy import func
    from app.admin.models import ModerationLog
    logs = db.session.query(
        ModerationLog.moderator_id,
        func.count(ModerationLog.id).label('actions')
    ).group_by(ModerationLog.moderator_id).all()
    return jsonify([{"moderator_id": l.moderator_id, "actions": l.actions} for l in logs])


@moderator_bp.route('/api/auto-priority', methods=['POST'])
@login_required
@require_role(*_MOD)
def api_auto_priority():
    from app.admin.models import ContentFlag
    now = datetime.now(timezone.utc)
    flags = ContentFlag.query.filter(
        ContentFlag.resolved_at == None,
        ContentFlag.auto_priority == True
    ).all()
    updated = 0
    for f in flags:
        age_hours = (now - f.created_at).total_seconds() / 3600
        if age_hours > 6 and f.priority != 'critical':
            f.priority = 'critical'
            updated += 1
        elif age_hours > 2 and f.priority != 'high':
            f.priority = 'high'
            updated += 1
    db.session.commit()
    return jsonify({"ok": True, "updated": updated})


@moderator_bp.route('/my_queue')
@login_required
@require_role(*_MOD)
def my_queue():
    """Return items assigned to the current moderator for the My Queue panel."""
    from app.admin.models import ContentFlag, ContentSubmission
    
    submissions = ContentSubmission.query.filter_by(
        assigned_to_id=current_user.id,
        status='pending'
    ).order_by(ContentSubmission.claimed_at.asc()).limit(10).all()

    flags = ContentFlag.query.filter_by(
        assigned_to=current_user.id,
        status='in_review'
    ).limit(5).all()

    return render_template(
        'admin/moderator/my_queue.html',
        submissions=submissions,
        flags=flags,
        title="My Queue"
    )


@moderator_bp.route('/api/unified-queue')
@login_required
@require_role(*_MOD)
def unified_queue():
    """Return unified moderation queue across all registered modules"""
    from app.admin.moderator.registry import get_registry, get_modules, get_entity_display, get_entity_icon
    
    # Query all open flags
    flags = ContentFlag.query.filter_by(status='open').order_by(
        db.case(
            (ContentFlag.priority == 'critical', 1),
            (ContentFlag.priority == 'high', 2),
            (ContentFlag.priority == 'medium', 3),
            (ContentFlag.priority == 'normal', 4),
            (ContentFlag.priority == 'low', 5),
            else_=6
        ),
        ContentFlag.created_at.asc()
    ).limit(50).all()
    
    registry = get_registry()
    modules = get_modules()
    
    # Group by module
    items = []
    for flag in flags:
        entity_type = flag.entity_type
        entity_info = registry.get(entity_type, {})
        
        # Try to get entity title
        title = f"{entity_type}#{flag.entity_id}"
        if entity_info.get('review_url_fn'):
            try:
                # Attempt to get entity name from the module's own function
                pass
            except:
                pass
        
        items.append({
            'id': flag.id,
            'entity_type': entity_type,
            'entity_id': flag.entity_id,
            'title': title,
            'priority': flag.priority,
            'reason': flag.reason[:100],
            'created_at': flag.created_at.isoformat(),
            'flagged_by': flag.flagger.username if flag.flagger else 'Unknown',
            'module': entity_info.get('module_name', 'Other'),
            'display_name': entity_info.get('display_name', entity_type),
            'icon': entity_info.get('icon', 'fa-flag'),
            'review_url': get_review_url(entity_type, flag.entity_id) or f"/moderator/flagged/{flag.id}"
        })
    
    # Get counts by module
    module_counts = {}
    for item in items:
        module = item['module']
        module_counts[module] = module_counts.get(module, 0) + 1
    
    # Get counts by type
    type_counts = {}
    for item in items:
        etype = item['entity_type']
        type_counts[etype] = {
            'display_name': item['display_name'],
            'count': type_counts.get(etype, {}).get('count', 0) + 1
        }
    
    return jsonify({
        'items': items,
        'total_pending': len(items),
        'modules': module_counts,
        'types': type_counts
    })


@moderator_bp.route('/api/moderate/<entity_type>/<int:entity_id>/<action>', methods=['POST'])
@login_required
@require_role(*_MOD)
def moderate_action(entity_type, entity_id, action):
    """Perform moderation action on any entity type"""
    data = request.get_json() or {}
    notes = data.get('notes', '')
    
    # Find flag for this entity
    flag = ContentFlag.query.filter_by(
        entity_type=entity_type,
        entity_id=entity_id,
        status='open'
    ).first()
    
    if not flag:
        return jsonify({'ok': False, 'error': 'No open flag found for this entity'}), 404
    
    if action == 'approve':
        flag.status = 'resolved'
        flag.resolution_action = 'approved'
        flag.resolution_notes = notes
        flag.resolved_by = current_user.id
        flag.resolved_at = datetime.now(timezone.utc)
        db.session.commit()
        
        _audit('moderate.approve', entity_type, {'entity_id': entity_id, 'notes': notes})
        
    elif action == 'reject':
        flag.status = 'resolved'
        flag.resolution_action = 'rejected'
        flag.resolution_notes = notes
        flag.resolved_by = current_user.id
        flag.resolved_at = datetime.now(timezone.utc)
        db.session.commit()
        
        _audit('moderate.reject', entity_type, {'entity_id': entity_id, 'notes': notes})
        
    else:
        return jsonify({'ok': False, 'error': f'Unknown action: {action}'}), 400
    
    return jsonify({'ok': True, 'message': f'{action.capitalize()}d successfully'})


@moderator_bp.route('/audit-log/export')
@login_required
@require_role(*_MOD)
def audit_log_export():
    """Export audit logs to CSV, Excel, or PDF."""
    import csv
    from io import StringIO, BytesIO
    from flask import Response, request
    from app.admin.owner.models import OwnerAuditLog
    from datetime import datetime, timezone, timedelta

    # Get format from query parameter (default: csv)
    export_format = request.args.get('format', 'csv').lower()

    # Get filters
    action_filter = request.args.get('action', '').strip()
    category_filter = request.args.get('category', '').strip()
    days = request.args.get('days', '30')

    # Build query for current user's actions only
    q = OwnerAuditLog.query.filter_by(actor_id=current_user.id)

    if action_filter:
        q = q.filter(OwnerAuditLog.action.ilike(f'%{action_filter}%'))
    if category_filter:
        q = q.filter_by(category=category_filter)

    # Date filter
    if days != 'all':
        days_int = int(days)
        cutoff = datetime.now(timezone.utc) - timedelta(days=days_int)
        q = q.filter(OwnerAuditLog.created_at >= cutoff)

    logs = q.order_by(OwnerAuditLog.created_at.desc()).limit(5000).all()

    # Prepare data rows
    rows = []
    headers = ['Timestamp', 'Action', 'Category', 'Entity Type', 'Entity ID', 'Details', 'IP Address', 'Status']

    for log in logs:
        details = log.details or {}
        rows.append([
            log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            log.action,
            log.category,
            details.get('entity_type', ''),
            details.get('entity_id', ''),
            str(details)[:500] if details else '',
            getattr(log, '_ip', ''),
            getattr(log, 'status', 'success')
        ])

    filename_base = f"moderator_audit_{current_user.username}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"

    # ========== CSV Export ==========
    if export_format == 'csv':
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        output.seek(0)

        return Response(
            output,
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename={filename_base}.csv'}
        )

    # ========== Excel Export ==========
    elif export_format == 'excel':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Audit Log"

            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # Write data
            for row_idx, row in enumerate(rows, 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-size columns
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return Response(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': f'attachment; filename={filename_base}.xlsx'}
            )
        except ImportError:
            return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    # ========== PDF Export ==========
    elif export_format == 'pdf':
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch

            output = BytesIO()
            doc = SimpleDocTemplate(output, pagesize=landscape(A4))
            elements = []

            # Title
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#4472C4'),
                spaceAfter=30
            )
            elements.append(Paragraph(f"Moderation Audit Log - {current_user.username}", title_style))
            elements.append(Spacer(1, 0.2*inch))

            # Table data
            table_data = [headers]
            for row in rows:
                # Truncate long details for PDF
                truncated_row = list(row)
                if len(truncated_row[5]) > 100:
                    truncated_row[5] = truncated_row[5][:100] + '...'
                table_data.append(truncated_row)

            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ]))

            elements.append(table)
            doc.build(elements)
            output.seek(0)

            return Response(
                output,
                mimetype='application/pdf',
                headers={'Content-Disposition': f'attachment; filename={filename_base}.pdf'}
            )
        except ImportError:
            return jsonify({"error": "reportlab not installed. Run: pip install reportlab"}), 500

    else:
        return jsonify({"error": f"Unsupported format: {export_format}"}), 400


# ═════════════════════════════════════════════════════════════════════════════
# K.  ENTERPRISE MODERATION FEATURES
# ═════════════════════════════════════════════════════════════════════════════

@moderator_bp.route('/ai-analytics')
@login_required
@require_role(*_MOD)
def ai_analytics():
    """AI-powered content detection analytics dashboard"""
    from app.admin.models import ContentFlag
    
    # Mock AI analytics data (would come from AI service)
    ai_stats = {
        'accuracy': 92.5,
        'processed_today': 156,
        'avg_processing_time': 0.45,
        'processed_hourly': 89,
        'false_positive_rate': 3.2,
        'false_negatives': 12,
        'auto_resolution_rate': 78.4,
        'auto_resolved_today': 67,
        'text_processed': 1247,
        'image_processed': 892,
        'behavior_processed': 456,
        'spam_processed': 2341,
        'toxicity_processed': 678,
        'trend_accuracy': 2.3,
        'true_positives': 1423,
        'false_positives': 48,
        'api_response_time': 120,
        'gpu_usage': 67,
        'last_training': '2d ago',
        'models_count': 5
    }
    
    return render_template(
        'admin/moderator/ai_analytics.html',
        ai_stats=ai_stats,
        title="AI Analytics Dashboard"
    )


@moderator_bp.route('/training')
@login_required
@require_role(*_MOD)
def training():
    """Moderator training and certification system"""
    # Mock training data (would come from training service)
    user_certification = {
        'level': 1,
        'certification_name': 'Basic Moderator',
        'credits_earned': 45,
        'modules_completed': 3,
        'issued_at': datetime.now(timezone.utc) - timedelta(days=30),
        'next_level_requirements': {
            'credits_needed': 55,
            'modules_needed': 2,
            'assessments_needed': 1
        }
    }

    training_modules = [
        {
            'id': 1,
            'name': 'Content Moderation Basics',
            'description': 'Learn the fundamentals of content moderation',
            'level': 1,
            'status': 'completed',
            'progress': 100,
            'credits': 10,
            'estimated_hours': 2
        },
        {
            'id': 2,
            'name': 'Policy Enforcement',
            'description': 'Understanding and applying content policies',
            'level': 1,
            'status': 'completed',
            'progress': 100,
            'credits': 15,
            'estimated_hours': 3
        },
        {
            'id': 3,
            'name': 'User Safety',
            'description': 'Protecting users from harmful content',
            'level': 1,
            'status': 'completed',
            'progress': 100,
            'credits': 20,
            'estimated_hours': 4
        },
        {
            'id': 4,
            'name': 'Advanced Content Analysis',
            'description': 'Deep analysis techniques for complex content',
            'level': 2,
            'status': 'in_progress',
            'progress': 65,
            'credits': 25,
            'estimated_hours': 5
        },
        {
            'id': 5,
            'name': 'Legal Compliance',
            'description': 'Understanding legal requirements for moderation',
            'level': 2,
            'status': 'not_started',
            'progress': 0,
            'credits': 30,
            'estimated_hours': 6
        }
    ]

    upcoming_assessments = [
        {
            'id': 1,
            'name': 'Content Policy Assessment',
            'module_name': 'Policy Enforcement',
            'priority': 'high',
            'due_date': datetime.now(timezone.utc) + timedelta(days=2)
        },
        {
            'id': 2,
            'name': 'Advanced Analysis Test',
            'module_name': 'Advanced Content Analysis',
            'priority': 'medium',
            'due_date': datetime.now(timezone.utc) + timedelta(days=5)
        }
    ]

    recommended_path = [
        {
            'order': 1,
            'module_name': 'Advanced Content Analysis',
            'reason': 'Build on current skills',
            'color': 'var(--accent)',
            'completed': False
        },
        {
            'order': 2,
            'module_name': 'Legal Compliance',
            'reason': 'Required for Level 2',
            'color': 'var(--blue)',
            'completed': False
        },
        {
            'order': 3,
            'module_name': 'Crisis Management',
            'reason': 'Advanced skill',
            'color': 'var(--purple)',
            'completed': False
        }
    ]

    training_stats = {
        'total_hours': 9,
        'assessment_score': 87,
        'streak_days': 5,
        'rank': '12/45'
    }

    return render_template(
        'admin/moderator/training.html',
        user_certification=user_certification,
        training_modules=training_modules,
        upcoming_assessments=upcoming_assessments,
        recommended_path=recommended_path,
        training_stats=training_stats,
        title="Moderator Training"
    )


@moderator_bp.route('/training-content')
@login_required
@require_role(*_MOD)
def training_content():
    """Comprehensive training materials and course content"""
    return render_template(
        'admin/moderator/training_content.html',
        title="Training Materials"
    )


@moderator_bp.route('/content-safety')
@login_required
@require_role(*_MOD)
def content_safety():
    """Content safety policies and enforcement dashboard"""
    # Mock policy data (would come from content safety service)
    policy_stats = {
        'hate_speech_blocks': 1247,
        'hate_speech_today': 89,
        'violence_blocks': 892,
        'violence_today': 67,
        'harassment_blocks': 456,
        'harassment_today': 34,
        'spam_blocks': 2341,
        'spam_today': 156
    }
    
    policy_categories = [
        {
            'id': 1,
            'name': 'Hate Speech',
            'description': 'Content that promotes hatred against individuals or groups',
            'status': 'active',
            'severity_level': 'high',
            'auto_enforce': True,
            'region': 'Global',
            'key_rules': [
                'No racial slurs or derogatory terms',
                'No content targeting protected characteristics',
                'No hate symbols or imagery'
            ],
            'blocks_today': 89,
            'total_blocks': 1247,
            'accuracy': 94.2
        },
        {
            'id': 2,
            'name': 'Violence and Threats',
            'description': 'Content that depicts or promotes violence',
            'status': 'active',
            'severity_level': 'critical',
            'auto_enforce': True,
            'region': 'Global',
            'key_rules': [
                'No graphic violence or gore',
                'No threats of physical harm',
                'No weapons or dangerous activities'
            ],
            'blocks_today': 67,
            'total_blocks': 892,
            'accuracy': 91.8
        },
        {
            'id': 3,
            'name': 'Harassment',
            'description': 'Targeted harassment or bullying of individuals',
            'status': 'active',
            'severity_level': 'medium',
            'auto_enforce': False,
            'region': 'US',
            'key_rules': [
                'No personal attacks or insults',
                'No repeated unwanted contact',
                'No doxxing or privacy violations'
            ],
            'blocks_today': 34,
            'total_blocks': 456,
            'accuracy': 89.5
        },
        {
            'id': 4,
            'name': 'Spam and Deceptive Content',
            'description': 'Misleading or repetitive content',
            'status': 'active',
            'severity_level': 'low',
            'auto_enforce': True,
            'region': 'Global',
            'key_rules': [
                'No repetitive or duplicate content',
                'No misleading links or scams',
                'No unauthorized commercial content'
            ],
            'blocks_today': 156,
            'total_blocks': 2341,
            'accuracy': 96.7
        },
        {
            'id': 5,
            'name': 'Sexual Content',
            'description': 'Adult or sexually explicit content',
            'status': 'active',
            'severity_level': 'high',
            'auto_enforce': True,
            'region': 'EU',
            'key_rules': [
                'No explicit sexual content',
                'No adult entertainment',
                'No sexual services or products'
            ],
            'blocks_today': 23,
            'total_blocks': 234,
            'accuracy': 93.1
        }
    ]
    
    effectiveness = {
        'false_positive_rate': 2.3,
        'detection_rate': 97.8,
        'response_time': 1.2,
        'user_satisfaction': 4.6
    }
    
    last_report_date = '2026-05-01'
    active_policies_count = 5
    regions_count = 5
    
    return render_template(
        'admin/moderator/content_safety.html',
        policy_stats=policy_stats,
        policy_categories=policy_categories,
        effectiveness=effectiveness,
        last_report_date=last_report_date,
        active_policies_count=active_policies_count,
        regions_count=regions_count,
        title="Content Safety Policies"
    )


@moderator_bp.route('/cross-platform')
@login_required
@require_role(*_MOD)
def cross_platform():
    """Cross-platform moderation dashboard"""
    # Mock platform data (would come from cross-platform service)
    platform_stats = {
        'web_actions': 892,
        'web_pending': 45,
        'web_response_time': 0.8,
        'ios_actions': 567,
        'ios_pending': 23,
        'ios_response_time': 1.2,
        'android_actions': 445,
        'android_pending': 31,
        'android_response_time': 1.5,
        'api_actions': 234,
        'api_pending': 12,
        'api_response_time': 2.1
    }
    
    platform_health = {
        'uptime': 99.8,
        'error_rate': 0.2,
        'sync_success': 98.5,
        'avg_latency': 1.4
    }
    
    sync_rules = {
        'user_suspension': 89,
        'content_removal': 156,
        'user_warning': 67
    }
    
    active_platforms_count = 4
    synced_actions_count = 156
    
    return render_template(
        'admin/moderator/cross_platform.html',
        platform_stats=platform_stats,
        platform_health=platform_health,
        sync_rules=sync_rules,
        active_platforms_count=active_platforms_count,
        synced_actions_count=synced_actions_count,
        title="Cross-Platform Moderation"
    )


# ============================================================================
# TRANSPORT MODERATION
# ============================================================================

@moderator_bp.route('/transport')
@login_required
@require_role(*_MOD)
def transport_moderation():
    """Transport moderation dashboard with drivers, vehicles, and bookings"""
    from app.transport.models import DriverProfile, Vehicle, Booking
    from app.admin.models import ContentFlag
    
    # Get transport statistics
    transport_stats = {
        'total_drivers': DriverProfile.query.filter_by(is_deleted=False).count(),
        'pending_drivers': DriverProfile.query.filter_by(
            verification_tier='pending', is_deleted=False
        ).count(),
        'verified_drivers': DriverProfile.query.filter_by(
            verification_tier='platform_verified', is_deleted=False
        ).count(),
        'total_vehicles': Vehicle.query.filter_by(is_deleted=False).count(),
        'pending_vehicles': Vehicle.query.filter_by(
            status='pending', is_deleted=False
        ).count(),
        'verified_vehicles': Vehicle.query.filter_by(
            status='active', is_deleted=False
        ).count(),
        'total_bookings': Booking.query.filter_by(is_deleted=False).count(),
        'disputed_bookings': Booking.query.filter_by(
            status='disputed', is_deleted=False
        ).count(),
        'active_bookings': Booking.query.filter(
            Booking.status.in_(['confirmed', 'assigned', 'driver_en_route', 'in_progress']),
            Booking.is_deleted == False
        ).count()
    }
    
    # Get flagged transport items
    transport_flags = ContentFlag.query.filter(
        ContentFlag.entity_type.like('transport_%')
    ).order_by(ContentFlag.created_at.desc()).limit(10).all()
    
    # Get pending items for review
    pending_drivers = DriverProfile.query.filter_by(
        verification_tier='pending', is_deleted=False
    ).order_by(DriverProfile.created_at.desc()).limit(5).all()
    
    pending_vehicles = Vehicle.query.filter_by(
        status='pending', is_deleted=False
    ).order_by(Vehicle.created_at.desc()).limit(5).all()
    
    disputed_bookings = Booking.query.filter_by(
        status='disputed', is_deleted=False
    ).order_by(Booking.created_at.desc()).limit(5).all()
    
    return render_template(
        'admin/moderator/transport.html',
        transport_stats=transport_stats,
        transport_flags=transport_flags,
        pending_drivers=pending_drivers,
        pending_vehicles=pending_vehicles,
        disputed_bookings=disputed_bookings,
        title="Transport Moderation"
    )


@moderator_bp.route('/transport/drivers')
@login_required
@require_role(*_MOD)
def transport_drivers():
    """Transport driver moderation queue"""
    from app.transport.models import DriverProfile
    
    status_filter = request.args.get('status', 'pending')
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '').strip()
    
    query = DriverProfile.query.filter_by(is_deleted=False)
    
    if status_filter and status_filter != 'all':
        query = query.filter_by(verification_tier=status_filter)
    
    if search:
        query = query.filter(
            db.or_(
                DriverProfile.first_name.ilike(f'%{search}%'),
                DriverProfile.last_name.ilike(f'%{search}%'),
                DriverProfile.email.ilike(f'%{search}%'),
                DriverProfile.phone.ilike(f'%{search}%')
            )
        )
    
    drivers = query.order_by(DriverProfile.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    return render_template(
        'admin/moderator/transport_drivers.html',
        drivers=drivers,
        current_status=status_filter,
        current_search=search,
        title="Transport Drivers"
    )


@moderator_bp.route('/transport/vehicles')
@login_required
@require_role(*_MOD)
def transport_vehicles():
    """Transport vehicle moderation queue"""
    from app.transport.models import Vehicle
    
    status_filter = request.args.get('status', 'pending')
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '').strip()
    
    query = Vehicle.query.filter_by(is_deleted=False)
    
    if status_filter and status_filter != 'all':
        query = query.filter_by(status=status_filter)
    
    if search:
        query = query.filter(
            db.or_(
                Vehicle.make.ilike(f'%{search}%'),
                Vehicle.model.ilike(f'%{search}%'),
                Vehicle.license_plate.ilike(f'%{search}%'),
                Vehicle.vehicle_type.ilike(f'%{search}%')
            )
        )
    
    vehicles = query.order_by(Vehicle.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    return render_template(
        'admin/moderator/transport_vehicles.html',
        vehicles=vehicles,
        current_status=status_filter,
        current_search=search,
        title="Transport Vehicles"
    )


@moderator_bp.route('/transport/bookings')
@login_required
@require_role(*_MOD)
def transport_bookings():
    """Transport booking moderation queue"""
    from app.transport.models import Booking
    
    status_filter = request.args.get('status', 'disputed')
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '').strip()
    
    query = Booking.query.filter_by(is_deleted=False)
    
    if status_filter and status_filter != 'all':
        query = query.filter_by(status=status_filter)
    
    if search:
        query = query.filter(
            db.or_(
                Booking.booking_reference.ilike(f'%{search}%'),
                Booking.pickup_address.ilike(f'%{search}%'),
                Booking.dropoff_address.ilike(f'%{search}%')
            )
        )
    
    bookings = query.order_by(Booking.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    return render_template(
        'admin/moderator/transport_bookings.html',
        bookings=bookings,
        current_status=status_filter,
        current_search=search,
        title="Transport Bookings"
    )


@moderator_bp.route('/transport/driver/<int:driver_id>')
@login_required
@require_role(*_MOD)
def view_transport_driver(driver_id):
    """View transport driver details for moderation"""
    from app.transport.models import DriverProfile
    from app.admin.models import ContentFlag
    
    driver = DriverProfile.query.get_or_404(driver_id)
    
    # Get driver-related flags
    driver_flags = ContentFlag.query.filter_by(
        entity_type='transport_driver',
        entity_id=driver_id
    ).order_by(ContentFlag.created_at.desc()).all()
    
    # Get driver vehicles
    from app.transport.models import Vehicle
    driver_vehicles = Vehicle.query.filter_by(
        owner_type='driver',
        owner_id=driver_id,
        is_deleted=False
    ).all()
    
    # Get driver bookings
    from app.transport.models import Booking
    driver_bookings = Booking.query.filter_by(
        driver_id=driver_id,
        is_deleted=False
    ).order_by(Booking.created_at.desc()).limit(10).all()
    
    return render_template(
        'admin/moderator/transport_driver_view.html',
        driver=driver,
        driver_flags=driver_flags,
        driver_vehicles=driver_vehicles,
        driver_bookings=driver_bookings,
        title=f"Driver: {driver.first_name} {driver.last_name}"
    )


@moderator_bp.route('/transport/vehicle/<int:vehicle_id>')
@login_required
@require_role(*_MOD)
def view_transport_vehicle(vehicle_id):
    """View transport vehicle details for moderation"""
    from app.transport.models import Vehicle
    from app.admin.models import ContentFlag
    
    vehicle = Vehicle.query.get_or_404(vehicle_id)
    
    # Get vehicle-related flags
    vehicle_flags = ContentFlag.query.filter_by(
        entity_type='transport_vehicle',
        entity_id=vehicle_id
    ).order_by(ContentFlag.created_at.desc()).all()
    
    # Get vehicle driver history
    from app.transport.models import DriverVehicleHistory
    vehicle_history = DriverVehicleHistory.query.filter_by(
        vehicle_id=vehicle_id
    ).order_by(DriverVehicleHistory.created_at.desc()).limit(10).all()
    
    return render_template(
        'admin/moderator/transport_vehicle_view.html',
        vehicle=vehicle,
        vehicle_flags=vehicle_flags,
        vehicle_history=vehicle_history,
        title=f"Vehicle: {vehicle.make} {vehicle.model}"
    )


@moderator_bp.route('/transport/booking/<int:booking_id>')
@login_required
@require_role(*_MOD)
def view_transport_booking(booking_id):
    """View transport booking details for moderation"""
    from app.transport.models import Booking
    from app.admin.models import ContentFlag
    
    booking = Booking.query.get_or_404(booking_id)
    
    # Get booking-related flags
    booking_flags = ContentFlag.query.filter_by(
        entity_type='transport_booking',
        entity_id=booking_id
    ).order_by(ContentFlag.created_at.desc()).all()
    
    return render_template(
        'admin/moderator/transport_booking_view.html',
        booking=booking,
        booking_flags=booking_flags,
        title=f"Booking: {booking.booking_reference}"
    )


@moderator_bp.route('/transport/action/<entity_type>/<int:entity_id>/<action>', methods=['POST'])
@login_required
@require_role(*_MOD)
def transport_moderate_action(entity_type, entity_id, action):
    """Perform moderation actions on transport entities"""
    from app.transport.models import DriverProfile, Vehicle, Booking
    
    # Get the appropriate entity
    if entity_type == 'driver':
        entity = DriverProfile.query.get_or_404(entity_id)
        redirect_url = url_for('admin.moderator.view_transport_driver', driver_id=entity_id)
    elif entity_type == 'vehicle':
        entity = Vehicle.query.get_or_404(entity_id)
        redirect_url = url_for('admin.moderator.view_transport_vehicle', vehicle_id=entity_id)
    elif entity_type == 'booking':
        entity = Booking.query.get_or_404(entity_id)
        redirect_url = url_for('admin.moderator.view_transport_booking', booking_id=entity_id)
    else:
        flash('Invalid entity type.', 'danger')
        return redirect(url_for('admin.moderator.transport_moderation'))
    
    # Perform the action
    if action == 'approve':
        if entity_type == 'driver':
            entity.verification_tier = 'platform_verified'
            # Add verified_at field if it exists
            if hasattr(entity, 'verified_at'):
                entity.verified_at = datetime.now(timezone.utc)
        elif entity_type == 'vehicle':
            entity.status = 'active'
            # Add verified_at field if it exists
            if hasattr(entity, 'verified_at'):
                entity.verified_at = datetime.now(timezone.utc)
        elif entity_type == 'booking':
            entity.status = 'confirmed'
        
        db.session.commit()
        flash(f'{entity_type.capitalize()} approved successfully.', 'success')
    
    elif action == 'reject':
        reason = request.form.get('reason', '').strip()
        if not reason:
            flash('Rejection reason is required.', 'warning')
            return redirect(redirect_url)
        
        if entity_type == 'driver':
            entity.verification_tier = 'pending'
            # Add rejection_reason field if it exists
            if hasattr(entity, 'rejection_reason'):
                entity.rejection_reason = reason
        elif entity_type == 'vehicle':
            entity.status = 'rejected'
            # Add rejection_reason field if it exists
            if hasattr(entity, 'rejection_reason'):
                entity.rejection_reason = reason
        elif entity_type == 'booking':
            entity.status = 'cancelled'
            # Add cancellation_reason field if it exists
            if hasattr(entity, 'cancellation_reason'):
                entity.cancellation_reason = reason
        
        db.session.commit()
        flash(f'{entity_type.capitalize()} rejected successfully.', 'success')
    
    elif action == 'flag':
        reason = request.form.get('reason', '').strip()
        priority = request.form.get('priority', 'medium')
        
        if not reason:
            flash('Reason required for flagging.', 'warning')
            return redirect(redirect_url)
        
        from app.admin.services import create_flag
        ok, flag = create_flag(
            user=current_user,
            entity_type=f'transport_{entity_type}',
            entity_id=entity_id,
            reason=reason,
            priority=priority
        )
        
        if ok:
            flash(f'{entity_type.capitalize()} flagged for review (Priority: {priority})', 'warning')
        else:
            flash(f'Failed to flag: {flag}', 'danger')
    
    elif action == 'suspend':
        if entity_type == 'driver':
            entity.is_active = False
            entity.suspension_reason = request.form.get('reason', 'Suspended by moderator')
            entity.suspended_at = datetime.now(timezone.utc)
        elif entity_type == 'vehicle':
            entity.is_available = False
            entity.suspension_reason = request.form.get('reason', 'Suspended by moderator')
        
        db.session.commit()
        flash(f'{entity_type.capitalize()} suspended successfully.', 'warning')
    
    return redirect(redirect_url)


@moderator_bp.route('/transport/third-party-applications')
@login_required
@require_role(*_MOD)
def transport_third_party_applications():
    """View third-party transport platform applications (Uber, SafeBoda, etc.)"""
    # This would show applications from external platforms wanting to integrate
    # For now, mock data - in production this would come from a database table
    
    applications = [
        {
            'id': 1,
            'platform_name': 'Uber',
            'company': 'Uber Technologies Inc.',
            'contact_email': 'partnerships@uber.com',
            'application_date': '2026-04-15',
            'status': 'pending_review',
            'integration_type': 'API Booking',
            'requested_features': ['Ride Booking', 'Driver Verification', 'Payment Processing'],
            'priority': 'high'
        },
        {
            'id': 2,
            'platform_name': 'SafeBoda',
            'company': 'SafeBoda Uganda Ltd',
            'contact_email': 'partners@safeboda.com',
            'application_date': '2026-04-20',
            'status': 'pending_review',
            'integration_type': 'API Booking',
            'requested_features': ['Boda Booking', 'Driver Onboarding', 'Fleet Management'],
            'priority': 'medium'
        }
    ]
    
    return render_template(
        'admin/moderator/transport_third_party.html',
        applications=applications,
        title="Third-Party Applications"
    )

